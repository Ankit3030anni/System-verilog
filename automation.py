
import sys
import re
import os
import subprocess
import re
import csv
from tabnanny import check
from textwrap import fill
from tkinter import SOLID
from turtle import color

from collections import OrderedDict
 

 


#os.system("pip install -r /home/pd_blr_imx91_blr01/users/ankit/anni/requirements.txt")

import openpyxl

from subprocess import Popen, PIPE


from openpyxl import Workbook, load_workbook

from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
    GradientFill,
)

#from openpyxl_image_loader import SheetImageLoader

n = len(sys.argv)
print(n)

command = ""

curr_dir = sys.argv[1]

print("the last value is:", curr_dir);
if (curr_dir[-1] == '/'):
  curr_dir = curr_dir[0:len(curr_dir)-1];


design_name = list(curr_dir.split("/"))[-1]
print("The TAG name is", design_name)
#print("type is ", type(design_name))

ppp2_file = subprocess.run('pwd', shell=True, capture_output=True, text=True);

#print(ppp2_file.stdout);

print(ppp2_file.stderr);
if (len(ppp2_file.stderr)!=0):
  sys.exit("MAKE SURE , YOU ARE IN YOUR DIR");
ppp2_file = ppp2_file.stdout.rstrip("\n");
#print(ppp2_file.stdout);
#sys.exit("DDoen");
print("Your current directory is",ppp2_file);
os.chdir(curr_dir)


des= f"grep 'set SVAR(design_name)' scripts_block/conf/block.tcl | head -1 | awk '{{print $3}}'";
file_des = subprocess.check_output(des, shell=True).decode("utf-8")

file_des = file_des.strip('"');
file_des = file_des.rstrip('"\n');

ppp2 = file_des;

#print(type(ppp2));
print("The design name is",file_des);


#ls 35_pnr/rpts/430_icc2_clock_opt_part1_qor/*.png


#Snap_img_list = []

#try:
    #Snaps = f"ls 35_pnr/rpts/*/*.png"
    #Snap_Img = subprocess.check_output(Snaps, shell=True).decode("utf-8")
    ##print("try")
    #Snap_img_list = Snap_Img.split("\n")
    #if Snap_Img == "":
        #Snap_img_list.append("")

#except:
    #Snap_img_list.append("")


##print("THe snap list is", Snap_img_list)
#if len(Snap_img_list) == 1:
    #print("NO, Image is Present")

#sheet_no = []
#total_stage_snap = [];
#for i in range(0, len(Snap_img_list) - 1):
    #a = list(Snap_img_list[i].split("/"))[2]
    #total_stage_snap.append(a);
    
#total_stage_snap = [*set(total_stage_snap)];

#for i in range(0, len(total_stage_snap)):
    #sheet_no.append(f"sheet{i+1}")

##print("the sheet_name is", sheet_no)
#A1 = ["A1", "AF1", "A52"];  #name of the image
#A2 = ["A2", "AF2", "A53"];   #location of the Image
#print("the ttal stage which has png are :",total_stage_snap);
#for i in range(0,len(total_stage_snap)):
  #a = f"ls 35_pnr/rpts/{total_stage_snap[i]}/*.png";
  #print(a);
  #stg_images = subprocess.run(a, shell=True, capture_output = True, text = True );
  #print (stg_images);
  #if (stg_images.stderr == ""):
    #stg_img_ans = stg_images.stdout.strip().split();
    #print("types us ",type(stg_img_ans))
    #print("the stge img ans is", stg_img_ans)
    #for j in range(0, len(stg_img_ans)):
      #name1 = stg_img_ans[j].split("/")[-2];
      #sheet_no[i] = work_obj.create_sheet(name1)
      #sheet_no[i][f'{A1[j]}'] = stg_img_ans[j].split("/")[-1];
      #img = Image(f'{sys.argv[1]}/{stg_img_ans[j]}');
      #img.anchor = A2[j];
      #sheet_no[i].add_image(img, img.anchor);
    

#for i in range(0, len(Snap_img_list) - 1):
    #name = list(Snap_img_list[i].split("/"))[2]
    #name1 = list(Snap_img_list[i].split("/"))[3]
    #print("The name type is", type(name))
    #name = f'35_pnr_{name}'
    #sheet_no[i] = work_obj.create_sheet(name1)
    #sheet_no[i].merge_cells("A1:D1")
    #sheet_no[i]["A1"] = name
    #cong_img = f"{sys.argv[1]}/{Snap_img_list[i]}"
    #img = Image(cong_img)
    #img.anchor = "A2"
    #sheet_no[i].add_image(img, img.anchor)


#print("\n\n")

#sys.exit("Image done")

#mm = f"grep 'TS_C16' 35_pnr/rpts/430_icc2_clock_opt_part1_qor/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"
#mm_o = subprocess.check_output(mm,stdin=subprocess.PIPE, shell = True).decode("utf-8");
#print("the mm_o is", mm_o);
#sys.exit("haha");


#std_struct_cl = f'grep "so std struct preserves not loaded to icc2 db" 25_dp/logs/750_icc2_dp_write_floorplan/icc2_dp_write_floorplan.log | grep "^SNPS"';
#std_struct = subprocess.run(
            #std_struct_cl, shell=True, capture_output = True, text = True
        #)
#print("the error is", std_struct.stderr)
#std_struct = std_struct.stdout.rstrip("\n");
#print("the value is",std_struct);

#sys.exit("done");
#try:
  #trail = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/430_icc2_clock_opt_part1_qor/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $5}}'";
  #stage_trail = subprocess.check_output(trail, stdin=subprocess.PIPE,shell=True).decode("utf-8")
  #print(stage_trail);
  #print("pqy");
  #print(stage_trail.stderr)
#except:
  #print("OK");

#grep_p_UI = f"grep 'SNPS_INFO   : Process for.*Preserve_UI.xlsx' 10_syn/rpts/020_dc_baseline/nxp_preserve_logic_csvs.rpt | awk '{{print $6}}'"
#p_UI = subprocess.check_output(grep_p_UI, stdin=subprocess.PIPE, shell=True).decode(
    #"utf-8")

#p_UI = p_UI.rstrip("\n")

#golden = "/home/pd_blr_imx95_blr03/users/devika/SOC_BE_IMX95_CLN16FFC_1.21.1.1/backend/common_flows/common_lynx_dft_flows/imx95_top/DOCUMENTS/Preserve_UI.xlsx"

#grep_f_p_UI = f'diff {p_UI} {golden}'
##l = f'diff /home/pd_blr_imx95_blr03/users/devika/SOC_BE_IMX95_CLN16FFC_1.20/backend/common_flows/common_lynx_dft_flows/imx95_top/DOCUMENTS/Preserve_UI.xlsx /home/pd_blr_imx95_blr03/users/devika/SOC_BE_IMX95_CLN16FFC_1.21.1.1/backend/common_flows/common_lynx_dft_flows/imx95_top/DOCUMENTS/Preserve_UI.xlsx'
#preserve_UI = subprocess.run(grep_f_p_UI, shell=True, capture_output=True, text=True)
#print(preserve_UI.stdout)
#preserve_UI = preserve_UI.stdout.rstrip("\n")
#print("The preserve_UI is", preserve_UI);
#if preserve_UI == "":
    #preserve_UI = "PASS"
#else:
    #preserve_UI = "FAIL"


    
##stage_std_struct.append(std_struct);
#print("The preserve ////ui is", preserve_UI)
##stage_preserve.append(preserve_UI);


#sys.exit("done");



#-------------------------------------------------------------------------------------
is_Syn_error = False;
try:
    grep_file_syn = "ls 10_syn/rpts/*/dc.report_clock_gating"
    file_name_syn = subprocess.check_output(grep_file_syn, shell=True).decode("utf-8")
except:
    print("ERROR in the SYNTHESIS Flow (10_syn)_ please check")
    is_Syn_error = True;


if is_Syn_error == False:
  stage_name_syn = file_name_syn.split("\n")
  #print("The Synthesis stage name is ", stage_name_syn)
  
  
  
  
is_csv = True;

try:
  grep_csv_syn = 'ls 10_syn/rpts/*/synth_checks.csv';
  csv_syn = subprocess.check_output(grep_csv_syn, shell =True).decode("utf-8");
except:
  print("NO CSV file");
  is_csv = False;

if is_csv == True:
  csv_syn_li = csv_syn.split("\n")
  #print("The CSV is",csv_syn_li);



#--------------------------------------------------------------------------------
stage_name = [];
is_35pnr_error = False;

try:
    grep_file = "ls 35_pnr/rpts/*/icc2.report_utilization"
    file_name = subprocess.check_output(grep_file, shell=True).decode("utf-8")
except:
    is_35pnr_error = True;
    print("The 35_pnr is not present");

if is_35pnr_error == False:
  stage_name = file_name.split("\n")
  #print("The stage name is ", stage_name)


###########################################################################################################3

design_name_li = []
d_name = []

h_name = []
v_name = []

stage_util = []
stage_h = []
stage_v = []

stage_std_count = []
stage_std_cell_area = []

stage_AOB = []

stage_shorts = []
stage_both_dir = []
stage_overall = []
stage_runtime = []

stage_DRC = []
stage_ccd = []
stage_bbox = []

stage_REG2REG = []
stage_REG2IN = []
stage_REG2OUT = []

stage_REG2REG_72v_125C = []
stage_REG2IN_72v_125C = []
stage_REG2OUT_72v_125C = []

stage_REG2REG_81v_125C = []
stage_REG2IN_81v_125C = []
stage_REG2OUT_81v_125C = []
stage_LVT = []
stage_ULVT = [];
stage_SVT = [];

stage_LVT16 = []
stage_ULVT16 = [];
stage_SVT16 = [];

stage_LVT18 = []
stage_ULVT18 = [];
stage_SVT18 = [];

stage_LVT20 = []
stage_ULVT20 = [];
stage_SVT20 = [];

stage_LVT24 = []
stage_ULVT24 = [];
stage_SVT24 = [];


Des_Name_li = [];
log_ext = []
if is_35pnr_error == False:
  rpt = list(stage_name[0].split("/"))[0]
  log = list(stage_name[0].split("/"))[0]
  #Des_Name_li = [];

  for i in range(0, len(stage_name) - 1):
      Des_Name_li.append(file_des);
      design_name_li.append(design_name)
      d_name.append(list(stage_name[i].split("/"))[2])

  #print("The d_name is", d_name)

  #log_ext = []
  for s in d_name:
      log_ext.append(s[s.index("_") + 1 : len(s)])

  #print("THe log extension is  ", log_ext)

#-------------------------------------------------------------------------------------

Des_Name_li_syn = [];
design_name_li_syn = [];
d_name_syn = [];

stage_FE_Bom_syn = [];


try:
  FE_BOM_path = f"grep 'set.*SOC_FE_WORKSPACE' scripts_block/conf/define_rtl_design.tcl | awk '{{print $NF}}'";
  FE_BOM = subprocess.check_output(FE_BOM_path, stdin=subprocess.PIPE, shell=True).decode("utf-8")
  if (FE_BOM == ""):
    FE_BOM = "Not Found";
    FE_BOM_1 = "Not Found";
  else:
    FE_BOM = FE_BOM.lstrip('"');
    FE_BOM = FE_BOM.lstrip("'");
    FE_BOM = FE_BOM.rstrip('"\n');
    FE_BOM = FE_BOM.rstrip("'\n");
    FE_BOM = FE_BOM.rstrip('\n');
    
    
    #print("The FE_BOM is", FE_BOM)
    F_B = f"grep DPDM_Part_Number {FE_BOM}/.pi_md/pi.bom | grep -oE 'SOC.*'" #grep DPDM_Part_Number /home/pd_blr_imx95_blr03/users/devika/SOC_BE_IMX95_CLN16FFC_1.19/SOC_FE_IMX95_CLN16FFC_1.87.1.4/.pi_md/pi.bom | grep -oE 'SOC.*'
    #print("The F_B is", F_B);
    FE_BOM_1 = subprocess.check_output(F_B, stdin=subprocess.PIPE, shell=True).decode("utf-8")
    #print("THE FE_bom_1 IS", FE_BOM_1)
    if (FE_BOM_1 == ""):
      FE_BOM_1 = "Not Found";
    else :
      FE_BOM_1 = FE_BOM_1.rstrip("\n");
except:
  FE_BOM_1 = "Not Found";
  

if is_Syn_error == False:
  #Des_Name_li_syn = [];
  #design_name_li_syn = [];
  #Des_Name_li_syn = Des_Name_li;
  #d_name_syn = [];
  #print("The stage_name synthesis is",stage_name_syn)
  for i in range(0, len(stage_name_syn) - 1):
      Des_Name_li_syn.append(file_des);
      design_name_li_syn.append(design_name)
      stage_FE_Bom_syn.append(FE_BOM_1);
      d_name_syn.append(list(stage_name_syn[i].split("/"))[2])


  #print(Des_Name_li_syn);
  #print(design_name_li_syn);
  #print("The Synthsis d_name is", d_name_syn)
  #sys.exit("ok")

  #-------------------------------------------------------------------------------------------
log_ext_syn = []
if is_Syn_error == False:
  
  for s in d_name_syn:
      log_ext_syn.append(s[s.index("_") + 1 : len(s)])

  #print("THe log extension for syn is  ", log_ext_syn)



#--------------------------------------------------------------------

#25_dp in Checklist
grep_check_file = f'ls 25_dp/logs/[0-9]*/*.log'
check_file = subprocess.run(grep_check_file, capture_output = True, shell = True, text =True)
check_stage_name = [];  

log_ext_check = [];
icc_golden_ver = [];
stage_icc_ver = [];
stage_icc_ver_status = [];

is_NO_DP_flow = "False";
if (check_file.stderr== ""):
    check_file = check_file.stdout.rstrip("\n");
    check_file = check_file.split("\n");
    #print("the check file is", check_file);
    for i in range(0, len(check_file)):
        check_stage_name.append(list(check_file[i].split("/"))[2])
    
    #print("DP stage_name", check_stage_name);
    print("\n\n")
    for s in check_stage_name:
      log_ext_check.append(s[s.index("_") + 1 : len(s)])

    for i in range(0,len(check_stage_name)):

        if(check_stage_name[i] == "330_icc2_dp_create_pg_mesh" or check_stage_name[i] == "325_icc2_dp_create_physical_only"):
            icc_golden_ver.append("synopsys-icc2-/S-2021.06-SP5-1-T-20230315");
        else:
            icc_golden_ver.append("synopsys-icc2-/S-2021.06-SP5-VAL-20220622")
        grep_icc_ver = f"grep '^SNPS_INFO' 25_dp/logs/{check_stage_name[i]}/{log_ext_check[i]}.log | grep 'SEV(ver_icc2)' | awk '{{print $NF}}'";
        icc_ver = subprocess.run(grep_icc_ver, shell = True, capture_output= True, text = True);
        if (icc_ver.stderr == ""):
            icc_ver = icc_ver.stdout.rstrip('"\n');
            #print("the icc_version is",icc_ver)
            #icc_ver = icc_ver.rstrip('"');
            icc_ver = icc_ver.lstrip('"');
            if (icc_ver == icc_golden_ver[i]):
                icc_ver_status = "PASS";
            else:
                icc_ver_status = "FAIL";
        else:
            icc_ver = "NA";
            icc_ver_status = "FAIL";

        stage_icc_ver.append(icc_ver);
        stage_icc_ver_status.append(icc_ver_status);



else:
    print("NO DP FLOW");
    is_NO_DP_flow = "True";


  #---------------------------------

stage_gate_reg = []
stage_REG2REG_syn = []
stage_REG2OUT_syn = []
stage_REG2IN_syn = []

stage_REG2REG_72v_125C_syn = []
stage_REG2OUT_72v_125C_syn = []
stage_REG2IN_72v_125C_syn = []

stage_REG2REG_81v_125C_syn = []
stage_REG2OUT_81v_125C_syn = []
stage_REG2IN_81v_125C_syn = []



stage_version_syn = []
stage_1_bit_syn = []
stage_2_bit_syn = []
stage_4_bit_syn = []
stage_8_bit_syn = []
stage_Total_bit = []
stage_dft_trace = []
stage_port_buf = []
if is_Syn_error == False:
    #stage_gate_reg = []
    #stage_REG2REG_syn = []
    #stage_REG2OUT_syn = []
    #stage_REG2IN_syn = []
    #stage_version_syn = []
    #stage_1_bit_syn = []
    #stage_2_bit_syn = []
    #stage_4_bit_syn = []
    #stage_8_bit_syn = []
    #stage_Total_bit = []
    for i in range(0, len(stage_name_syn) - 1):
        gate_reg_syn = f"grep 'Number of gated registers' 10_syn/rpts/{d_name_syn[i]}/dc.report_clock_gating | awk '{{print $(NF-1)}}' | sed -r 's,\(|\),,g'"
        wns_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $5}}'"
        tns_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $7}}'"
        nve_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $9}}'"

        wns_REGIN_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $5}}'"
        tns_REGIN_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $7}}'"
        nve_REGIN_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $9}}'"

        wns_REGOUT_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $5}}'"
        tns_REGOUT_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $7}}'"
        nve_REGOUT_syn = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $9}}'"
        
        wns_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $5}}'"
        tns_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $7}}'"
        nve_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $9}}'"

        wns_REGIN_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $5}}'"
        tns_REGIN_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $7}}'"
        nve_REGIN_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $9}}'"

        wns_REGOUT_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $5}}'"
        tns_REGOUT_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $7}}'"
        nve_REGOUT_72v_125C_syn = f"grep FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $9}}'"

        wns_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $5}}'"
        tns_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $7}}'"
        nve_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $9}}'"

        wns_REGIN_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $5}}'"
        tns_REGIN_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $7}}'"
        nve_REGIN_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $9}}'"

        wns_REGOUT_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $5}}'"
        tns_REGOUT_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $7}}'"
        nve_REGOUT_81v_125C_syn = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 10_syn/rpts/{d_name_syn[i]}/dc.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $9}}'"

        
        
        grep_1_bit = f"grep '1-bit' 10_syn/rpts/{d_name_syn[i]}/dc.report_clock_gating | tail -1 | awk '{{print $6}}'"
        grep_2_bit = f"grep '2-bit' 10_syn/rpts/{d_name_syn[i]}/dc.report_clock_gating | tail -1 | awk '{{print $6}}'"
        grep_4_bit = f"grep '4-bit' 10_syn/rpts/{d_name_syn[i]}/dc.report_clock_gating | tail -1 | awk '{{print $6}}'"
        grep_8_bit = f"grep '8-bit' 10_syn/rpts/{d_name_syn[i]}/dc.report_clock_gating | tail -1 | awk '{{print $6}}'"

        grep_Total_bit = f"grep 'Total' 10_syn/rpts/{d_name_syn[i]}/dc.report_clock_gating | grep -v 'Total number of registers' | grep -v 'Total Number of Registers' | tail -1 | awk '{{print $6}}'"

        ver_syn = f"grep Version 10_syn/logs/{d_name_syn[i]}/{log_ext_syn[i]}.log | head -1 | awk '{{print $2}}'"
        
        dft_trace = f"egrep 'Traced.*Scan Elements' 10_syn/rpts/{d_name_syn[i]}/nxp_dft_insertion.rpt | tail -1 | awk '{{print $NF}}' | sed -r 's,\(|\),,g'"
        
        grep_port_buf = f"egrep '^SNPS.*port_buffer' 10_syn/logs/{d_name_syn[i]}/{log_ext_syn[i]}.log | awk '{{print $NF}}'"
        
        port_buf = subprocess.run(grep_port_buf, shell=True, capture_output=True, text=True);
        if (port_buf.stderr == ""):
            port_buf= port_buf.stdout.rstrip("\n");
            if (port_buf == "0"):
                port_buf = "Disabled";
            else:
                port_buf = "Enabled"
        else:
            port_buf = "Disabled";
        try:
            dft = subprocess.check_output(
                dft_trace, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            dft = dft.rstrip("\n")
            if dft == "":
                dft = "FAIL"
        except:
            dft = "FAIL"
        
        try:
            bit_1 = subprocess.check_output(
                grep_1_bit, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            bit_1 = bit_1.rstrip("\n")
            if bit_1 == "":
                bit_1 = "NA"
        except:
            bit_1 = "NA"

        try:
            bit_2 = subprocess.check_output(
                grep_2_bit, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            bit_2 = bit_2.rstrip("\n")
            if bit_2 == "":
                bit_2 = "NA"
        except:
            bit_2 = "NA"

        try:
            bit_4 = subprocess.check_output(
                grep_4_bit, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            bit_4 = bit_4.rstrip("\n")
            if bit_4 == "":
                bit_4 = "NA"
        except:
            bit_4 = "NA"

        try:
            bit_8 = subprocess.check_output(
                grep_8_bit, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            bit_8 = bit_8.rstrip("\n")
            if bit_8 == "":
                bit_8 = "NA"
        except:
            bit_8 = "NA"

        try:
            Total_bit = subprocess.check_output(
                grep_Total_bit, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            Total_bit = Total_bit.rstrip("\n")
            if Total_bit == "":
                Total_bit = "NA"
        except:
            Total_bit = "NA"

        try:
            gate_reg = subprocess.check_output(
                gate_reg_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            gate_reg = gate_reg.rstrip("\n")
            if gate_reg == "":
                gate_reg = "NA"
        except:
            gate_reg = "NA"
        try:
            ans1_wns = subprocess.check_output(
                wns_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_wns = ans1_wns.rstrip("\n")
            if ans1_wns == "":
                ans1_wns = "NA"
        except:
            ans1_wns = "NA"
        try:
            ans1_tns = subprocess.check_output(
                tns_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_tns = ans1_tns.rstrip("\n")
            if ans1_tns == "":
                ans1_tns = "NA"
        except:
            ans1_tns = "NA"

        try:
            ans1_nve = subprocess.check_output(
                nve_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_nve = ans1_nve.rstrip("\n")
            if ans1_nve == "":
                ans1_nve = "NA"
        except:
            ans1_nve = "NA"

        try:
            ans1_REGIN_wns = subprocess.check_output(
                wns_REGIN_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_REGIN_wns = ans1_REGIN_wns.rstrip("\n")
            if ans1_REGIN_wns == "":
                ans1_REGIN_wns = "NA"
        except:
            ans1_REGIN_wns = "NA"

        try:
            ans1_REGIN_tns = subprocess.check_output(
                tns_REGIN_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_REGIN_tns = ans1_REGIN_tns.rstrip("\n")
            if ans1_REGIN_tns == "":
                ans1_REGIN_tns = "NA"
        except:
            ans1_REGIN_tns = "NA"

        try:
            ans1_REGIN_nve = subprocess.check_output(
                nve_REGIN_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_REGIN_nve = ans1_REGIN_nve.rstrip("\n")
            if ans1_REGIN_nve == "":
                ans1_REGIN_nve = "NA"
        except:
            ans1_REGIN_nve = "NA"

        try:
            ans1_REGOUT_wns = subprocess.check_output(
                wns_REGOUT_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_REGOUT_wns = ans1_REGOUT_wns.rstrip("\n")
            if ans1_REGOUT_wns == "":
                ans1_REGOUT_wns = "NA"
        except:
            ans1_REGOUT_wns = "NA"

        try:
            ans1_REGOUT_tns = subprocess.check_output(
                tns_REGOUT_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_REGOUT_tns = ans1_REGOUT_tns.rstrip("\n")
            if ans1_REGOUT_tns == "":
                ans1_REGOUT_tns = "NA"
        except:
            ans1_REGOUT_tns = "NA"

        try:
            ans1_REGOUT_nve = subprocess.check_output(
                nve_REGOUT_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_REGOUT_nve = ans1_REGOUT_nve.rstrip("\n")
            if ans1_REGOUT_nve == "":
                ans1_REGOUT_nve = "NA"
        except:
            ans1_REGOUT_nve = "NA"
            
            
        try:
            ans1_72v_125C_wns = subprocess.check_output(
                wns_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_wns = ans1_72v_125C_wns.rstrip("\n")
            if ans1_72v_125C_wns == "":
                ans1_72v_125C_wns = "NA"
        except:
            ans1_72v_125C_wns = "NA"
        try:
            ans1_72v_125C_tns = subprocess.check_output(
                tns_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_tns = ans1_72v_125C_tns.rstrip("\n")
            if ans1_72v_125C_tns == "":
                ans1_72v_125C_tns = "NA"
        except:
            ans1_72v_125C_tns = "NA"

        try:
            ans1_72v_125C_nve = subprocess.check_output(
                nve_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_nve = ans1_72v_125C_nve.rstrip("\n")
            if ans1_72v_125C_nve == "":
                ans1_72v_125C_nve = "NA"
        except:
            ans1_72v_125C_nve = "NA"

        try:
            ans1_72v_125C_REGIN_wns = subprocess.check_output(
                wns_REGIN_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_REGIN_wns = ans1_72v_125C_REGIN_wns.rstrip("\n")
            if ans1_72v_125C_REGIN_wns == "":
                ans1_72v_125C_REGIN_wns = "NA"
        except:
            ans1_72v_125C_REGIN_wns = "NA"

        try:
            ans1_72v_125C_REGIN_tns = subprocess.check_output(
                tns_REGIN_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_REGIN_tns = ans1_72v_125C_REGIN_tns.rstrip("\n")
            if ans1_72v_125C_REGIN_tns == "":
                ans1_72v_125C_REGIN_tns = "NA"
        except:
            ans1_72v_125C_REGIN_tns = "NA"

        try:
            ans1_72v_125C_REGIN_nve = subprocess.check_output(
                nve_REGIN_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_REGIN_nve = ans1_72v_125C_REGIN_nve.rstrip("\n")
            if ans1_72v_125C_REGIN_nve == "":
                ans1_72v_125C_REGIN_nve = "NA"
        except:
            ans1_72v_125C_REGIN_nve = "NA"

        try:
            ans1_72v_125C_REGOUT_wns = subprocess.check_output(
                wns_REGOUT_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_REGOUT_wns = ans1_72v_125C_REGOUT_wns.rstrip("\n")
            if ans1_72v_125C_REGOUT_wns == "":
                ans1_72v_125C_REGOUT_wns = "NA"
        except:
            ans1_72v_125C_REGOUT_wns = "NA"

        try:
            ans1_72v_125C_REGOUT_tns = subprocess.check_output(
                tns_REGOUT_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_REGOUT_tns = ans1_72v_125C_REGOUT_tns.rstrip("\n")
            if ans1_72v_125C_REGOUT_tns == "":
                ans1_72v_125C_REGOUT_tns = "NA"
        except:
            ans1_72v_125C_REGOUT_tns = "NA"

        try:
            ans1_72v_125C_REGOUT_nve = subprocess.check_output(
                nve_REGOUT_72v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_72v_125C_REGOUT_nve = ans1_72v_125C_REGOUT_nve.rstrip("\n")
            if ans1_72v_125C_REGOUT_nve == "":
                ans1_72v_125C_REGOUT_nve = "NA"
        except:
            ans1_72v_125C_REGOUT_nve = "NA"
            
        try:
            ans1_81_125C_wns = subprocess.check_output(
                wns_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_wns = ans1_81_125C_wns.rstrip("\n")
            #print("TTTTTTTTTTTTT", ans1_81_125C_wns)
            if ans1_81_125C_wns == "":
                ans1_81_125C_wns = "NA"
        except:
            ans1_81_125C_wns = "NA"
        try:
            ans1_81_125C_tns = subprocess.check_output(
                tns_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_tns = ans1_81_125C_tns.rstrip("\n")
            if ans1_81_125C_tns == "":
                ans1_81_125C_tns = "NA"
        except:
            ans1_81_125C_tns = "NA"

        try:
            ans1_81_125C_nve = subprocess.check_output(
                nve_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_nve = ans1_81_125C_nve.rstrip("\n")
            if ans1_81_125C_nve == "":
                ans1_81_125C_nve = "NA"
        except:
            ans1_81_125C_nve = "NA"

        try:
            ans1_81_125C_REGIN_wns = subprocess.check_output(
                wns_REGIN_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_REGIN_wns = ans1_81_125C_REGIN_wns.rstrip("\n")
            if ans1_81_125C_REGIN_wns == "":
                ans1_81_125C_REGIN_wns = "NA"
        except:
            ans1_81_125C_REGIN_wns = "NA"

        try:
            ans1_81_125C_REGIN_tns = subprocess.check_output(
                tns_REGIN_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_REGIN_tns = ans1_81_125C_REGIN_tns.rstrip("\n")
            if ans1_81_125C_REGIN_tns == "":
                ans1_81_125C_REGIN_tns = "NA"
        except:
            ans1_81_125C_REGIN_tns = "NA"

        try:
            ans1_81_125C_REGIN_nve = subprocess.check_output(
                nve_REGIN_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_REGIN_nve = ans1_81_125C_REGIN_nve.rstrip("\n")
            if ans1_81_125C_REGIN_nve == "":
                ans1_81_125C_REGIN_nve = "NA"
        except:
            ans1_81_125C_REGIN_nve = "NA"

        try:
            ans1_81_125C_REGOUT_wns = subprocess.check_output(
                wns_REGOUT_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_REGOUT_wns = ans1_81_125C_REGOUT_wns.rstrip("\n")
            if ans1_81_125C_REGOUT_wns == "":
                ans1_81_125C_REGOUT_wns = "NA"
        except:
            ans1_81_125C_REGOUT_wns = "NA"

        try:
            ans1_81_125C_REGOUT_tns = subprocess.check_output(
                tns_REGOUT_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_REGOUT_tns = ans1_81_125C_REGOUT_tns.rstrip("\n")
            if ans1_81_125C_REGOUT_tns == "":
                ans1_81_125C_REGOUT_tns = "NA"
        except:
            ans1_81_125C_REGOUT_tns = "NA"

        try:
            ans1_81_125C_REGOUT_nve = subprocess.check_output(
                nve_REGOUT_81v_125C_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans1_81_125C_REGOUT_nve = ans1_81_125C_REGOUT_nve.rstrip("\n")
            if ans1_81_125C_REGOUT_nve == "":
                ans1_81_125C_REGOUT_nve = "NA"
        except:
            ans1_81_125C_REGOUT_nve = "NA"


        try:
            ans2_version = subprocess.check_output(
                ver_syn, stdin=subprocess.PIPE, shell=True
            ).decode("utf-8")
            ans2_version = ans2_version.rstrip("\n")
            if ans2_version == "":
                ans2_version = "NA"
        except:
            ans2_version = "NA"
        stage_gate_reg.append(gate_reg)
        stage_REG2REG_syn.append(f"{ans1_wns} | {ans1_tns} | {ans1_nve}")
        stage_REG2IN_syn.append(
            f"{ans1_REGIN_wns} | {ans1_REGIN_tns} | {ans1_REGIN_nve}"
        )
        stage_REG2OUT_syn.append(
            f"{ans1_REGOUT_wns} | {ans1_REGOUT_tns} | {ans1_REGOUT_nve}"
        )
        stage_REG2REG_81v_125C_syn.append(f"{ans1_81_125C_wns} | {ans1_81_125C_tns} | {ans1_81_125C_nve}")
        stage_REG2IN_81v_125C_syn.append(
            f"{ans1_81_125C_REGIN_wns} | {ans1_81_125C_REGIN_tns} | {ans1_81_125C_REGIN_nve}"
        )
        stage_REG2OUT_81v_125C_syn.append(
            f"{ans1_81_125C_REGOUT_wns} | {ans1_81_125C_REGOUT_tns} | {ans1_81_125C_REGOUT_nve}"
        )
        stage_REG2REG_72v_125C_syn.append(f"{ans1_72v_125C_wns} | {ans1_72v_125C_tns} | {ans1_72v_125C_nve}")
        stage_REG2IN_72v_125C_syn.append(
            f"{ans1_72v_125C_REGIN_wns} | {ans1_72v_125C_REGIN_tns} | {ans1_72v_125C_REGIN_nve}"
        )
        stage_REG2OUT_72v_125C_syn.append(
            f"{ans1_72v_125C_REGOUT_wns} | {ans1_72v_125C_REGOUT_tns} | {ans1_72v_125C_REGOUT_nve}"
        )

        stage_version_syn.append(ans2_version)
        if bit_1 != "NA" and Total_bit != "NA":
            bit_1 = int(bit_1) / int(Total_bit) * 100;
            bit_1 = int(bit_1 * 10 ** 2) / 10 ** 2
            bit_1 = str(bit_1)
        else:
            bit_1 = "NA"
        if bit_2 != "NA" and Total_bit != "NA":
            bit_2 = int(bit_2) / int(Total_bit) * 100;
            bit_2 = int(bit_2 * 10 ** 2) / 10 ** 2
            bit_2 = str(bit_2)
        else:
            bit_2 = "NA"
        if bit_4 != "NA" and Total_bit != "NA":
            bit_4 = int(bit_4) / int(Total_bit) * 100
            bit_4 = int(bit_4 * 10 ** 2) / 10 ** 2
            bit_4 = str(bit_4)
        else:
            bit_4 = "NA"
        if bit_8 != "NA" and Total_bit != "NA":
            bit_8 = int(bit_8) / int(Total_bit) * 100
            bit_8 = int(bit_8 * 10 ** 2) / 10 ** 2
            bit_8 = str(bit_8)
        else:
            bit_8 = "NA"
        stage_1_bit_syn.append(bit_1)
        stage_2_bit_syn.append(bit_2)
        stage_4_bit_syn.append(bit_4)
        stage_8_bit_syn.append(bit_8)
        stage_Total_bit.append(Total_bit)
        stage_dft_trace.append(dft);
        stage_port_buf.append(port_buf);

    #print("synthesis stage stage_REG2REG_syn", stage_REG2REG_syn)
    #print("synthesis stage stage_REGIN_syn", stage_REG2IN_syn)
    #print("synthesis stage stage_REGOUT_syn_syn", stage_REG2OUT_syn)
    #print("DC_version syn is,", stage_version_syn)
    #print("Total single bit equivlaent:", stage_Total_bit)
    #print("1_bit value is :", stage_1_bit_syn)


#print("synthesis stage gete list is", stage_gate_reg)
#--------------------------------------------------------------------------------------------------------------------------------------------
#for_chcklist
stage_preserve = [];
stage_std_struct = [];
stage_dp_grid = [];
is_checklist = False
if is_checklist == False:

    std_struct_cl = f'grep "so std struct preserves not loaded to icc2 db" 25_dp/logs/750_icc2_dp_write_floorplan/icc2_dp_write_floorplan.log | grep "^SNPS"'
    std_struct = subprocess.run(
        std_struct_cl, shell=True, capture_output=True, text=True
    )
    if std_struct.stderr == "":
        std_struct = std_struct.stdout.rstrip("\n")
        if std_struct == "":
            std_struct = "PASS"
        else:
            std_struct = "FAIL"
    else:
        std_struct = "FAIL"

    grep_p_UI = f"grep 'SNPS_INFO   : Process for.*Preserve_UI.xlsx' 10_syn/rpts/020_dc_baseline/nxp_preserve_logic_csvs.rpt | awk '{{print $6}}'"
    p_UI = subprocess.run(grep_p_UI, shell=True, capture_output=True, text=True)
    if len(p_UI.stderr) == 0:
        p_UI = p_UI.stdout.rstrip("\n")

        print("P_UI is :", p_UI)

        golden = "/home/pd_blr_imx95_blr03/users/devika/SOC_BE_IMX95_CLN16FFC_1.21.1.1/backend/common_flows/common_lynx_dft_flows/imx95_top/DOCUMENTS/Preserve_UI.xlsx"

        grep_f_p_UI = f"diff {p_UI} {golden}"
        # l = f'diff /home/pd_blr_imx95_blr03/users/devika/SOC_BE_IMX95_CLN16FFC_1.20/backend/common_flows/common_lynx_dft_flows/imx95_top/DOCUMENTS/Preserve_UI.xlsx /home/pd_blr_imx95_blr03/users/devika/SOC_BE_IMX95_CLN16FFC_1.21.1.1/backend/common_flows/common_lynx_dft_flows/imx95_top/DOCUMENTS/Preserve_UI.xlsx'
        preserve_UI = subprocess.run(
            grep_f_p_UI, shell=True, capture_output=True, text=True
        )
        # print(preserve_UI.stdout)
        # print("The preserve error is",preserve_UI.stderr);
        preserve_UI = preserve_UI.stdout.rstrip("\n")
        # print("The preserve_UI is", preserve_UI);
        if preserve_UI == "":
            preserve_UI = "PASS"
        else:
            preserve_UI = "FAIL"
    else:
        preserve_UI = "FAIL"
        
    try:
        dp_grid = ""
        grep_SVAR = f'grep "SVAR(dp,snap_grid,exclude_macro,lib_cell_pattern_list)" scripts_block/conf/block.tcl'
        dp_grid = subprocess.run(grep_SVAR, shell=True, capture_output=True, text=True)
        dp_grid = dp_grid.stdout.rstrip("\n")
        if dp_grid != "":
            dp_grid = "PASS"
        else:
            dp_grid = "FAIL"
    except:
        dp_grid = "FAIL"

    stage_std_struct.append(std_struct)
    # print("The preserve ////ui is", preserve_UI)
    stage_preserve.append(preserve_UI)
    stage_dp_grid.append(dp_grid)




#----------------------------------------------------------------------------------------------------------------------------------------------

for i in range(0, len(stage_name) - 1):
    stage_grep_util = f'grep "Utilization Ratio" {rpt}/rpts/{d_name[i]}/icc2.report_utilization | grep -E -o "([0-9]*\.?[0-9]*)"'
    stage_grep_both_dir = f'grep "Both Dirs" {rpt}/rpts/{d_name[i]}/icc2.report_congestion | tail -1 | grep -E -o "([0-9]*\.?[0-9]*)\s*%" | sed "s/%//"'
    stage_grep_h = f'grep "H routing" {rpt}/rpts/{d_name[i]}/icc2.report_congestion | tail -1 | grep -E -o "([0-9]*\.?[0-9]*)\s*%" | sed "s/%//"'
    stage_grep_v = f'grep "V routing" {rpt}/rpts/{d_name[i]}/icc2.report_congestion | tail -1 | grep -E -o "([0-9]*\.?[0-9]*)\s*%" | sed "s/%//"'

    stage_grep_std_count = f'grep "Standard cells" {rpt}/rpts/{d_name[i]}/icc2.report_design | head -1 | grep -oE "[0-9]+\s"'
    stage_grep_std_area = f'grep "Standard cells" {rpt}/rpts/{d_name[i]}/icc2.report_design | head -1 | grep -E -o "([0-9]*\.[0-9]*)"'

    stage_grep_AOB = f'grep "Always on cells" {rpt}/rpts/{d_name[i]}/icc2.report_design | head -1 | grep -oE "[0-9]+\s"'

    stage_grep_Shorts = (
        f'grep "Short" {rpt}/rpts/{d_name[i]}/icc2.report_design | grep -oE "[0-9]+"'
    )
    stage_grep_bbox = f'grep "Chip Area" {rpt}/rpts/{d_name[i]}/icc2.report_design | tail -1 | grep -oE "[0-9]+\.[0-9]+"'

    stage_grep_DRC = f'grep "TOTAL VIOLATION" {rpt}/rpts/{d_name[i]}/icc2.report_design | grep -oE "[0-9]+" | tail -1'
    stage_grep_ccd = f'grep -c "^CCD-Info" {log}/logs/{d_name[i]}/{log_ext[i]}.log'
    
    stage_grep_LVT = f'grep "^LVT Area %" 35_pnr/rpts/{d_name[i]}/*.csv | grep -oE "[0-9]+\.[0-9]+"'; #grep 'TS_C16' 35_pnr/rpts/430_icc2_clock_opt_part1_qor/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'
    stage_grep_SVT = f'grep "^SVT Area %" 35_pnr/rpts/{d_name[i]}/*.csv | grep -oE "[0-9]+\.[0-9]+"';
    stage_grep_ULVT = f'grep "^ULVT Area %" 35_pnr/rpts/{d_name[i]}/*.csv | grep -oE "[0-9]+\.[0-9]+"';

    stage_grep_LVT16 = f"grep 'TL_C16' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'";
    stage_grep_SVT16 = f"grep 'TS_C16' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'";
    stage_grep_ULVT16 = f"grep 'TUL_C16' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'";

    stage_grep_LVT18 = f"grep 'TL_C18' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'";
    stage_grep_SVT18 = f"grep 'TS_C18' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"
    stage_grep_ULVT18 = f"grep 'TUL_C18' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"
    
    stage_grep_LVT20 = f"grep 'TL_C20' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"
    stage_grep_SVT20 = f"grep 'TS_C20' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"
    stage_grep_ULVT20 = f"grep 'TUL_C20' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"

    stage_grep_LVT24 = f"grep 'TL_C24' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"
    stage_grep_SVT24 = f"grep 'TS_C24' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"
    stage_grep_ULVT24 = f"grep 'TUL_C24' 35_pnr/rpts/{d_name[i]}/icc2.report_design_vt.csv | awk '{{print $NF}}' | sed 's/%//'"

    stage_grep_wns = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $5}}'"
    stage_grep_tns = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $7}}'"
    stage_grep_nve = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $9}}'"

    stage_grep_REGIN_wns = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $5}}'"
    stage_grep_REGIN_tns = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $7}}'"
    stage_grep_REGIN_nve = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $9}}'"

    stage_grep_REGOUT_wns = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $5}}'"
    stage_grep_REGOUT_tns = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $7}}'"
    stage_grep_REGOUT_nve = f"grep FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $9}}'"
    
    stage_grep_wns_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $5}}'"
    stage_grep_tns_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $7}}'"
    stage_grep_nve_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $9}}'"

    stage_grep_REGIN_wns_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $5}}'"
    stage_grep_REGIN_tns_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $7}}'"
    stage_grep_REGIN_nve_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $9}}'"

    stage_grep_REGOUT_wns_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $5}}'"
    stage_grep_REGOUT_tns_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $7}}'"
    stage_grep_REGOUT_nve_72v_125C = f"grep FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $9}}'"
    
    stage_grep_wns_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $5}}'"
    stage_grep_tns_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $7}}'"
    stage_grep_nve_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{{print $9}}'"

    stage_grep_REGIN_wns_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $5}}'"
    stage_grep_REGIN_tns_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $7}}'"
    stage_grep_REGIN_nve_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGIN | awk '{{print $9}}'"

    stage_grep_REGOUT_wns_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $5}}'"
    stage_grep_REGOUT_tns_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $7}}'"
    stage_grep_REGOUT_nve_81v_125C = f"grep FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REGOUT | awk '{{print $9}}'"
    #'grep "FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T" 35_pnr/rpts/{d_name[i]}/icc2.report_qor_reg2reg | grep -v TOTAL | grep REG2REG | awk '{print $9}''

    try:
        ans1_wns = subprocess.check_output(
            stage_grep_wns, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_wns = ans1_wns.rstrip("\n")
        if ans1_wns == "":
            ans1_wns = "NA"
    except:
        ans1_wns = "NA"

    try:
        ans1_tns = subprocess.check_output(
            stage_grep_tns, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_tns = ans1_tns.rstrip("\n")
        if ans1_tns == "":
            ans1_tns = "NA"
    except:
        ans1_tns = "NA"

    try:
        ans1_nve = subprocess.check_output(
            stage_grep_nve, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_nve = ans1_nve.rstrip("\n")
        if ans1_nve == "":
            ans1_nve = "NA"
    except:
        ans1_nve = "NA"

    try:
        ans1_REGIN_wns = subprocess.check_output(
            stage_grep_REGIN_wns, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_wns = ans1_REGIN_wns.rstrip("\n")
        if ans1_REGIN_wns == "":
            ans1_REGIN_wns = "NA"
    except:
        ans1_REGIN_wns = "NA"

    try:
        ans1_REGIN_tns = subprocess.check_output(
            stage_grep_REGIN_tns, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_tns = ans1_REGIN_tns.rstrip("\n")
        if ans1_REGIN_tns == "":
            ans1_REGIN_tns = "NA"
    except:
        ans1_REGIN_tns = "NA"

    try:
        ans1_REGIN_nve = subprocess.check_output(
            stage_grep_REGIN_nve, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_nve = ans1_REGIN_nve.rstrip("\n")
        if ans1_REGIN_nve == "":
            ans1_REGIN_nve = "NA"
    except:
        ans1_REGIN_nve = "NA"

    try:
        ans1_REGOUT_wns = subprocess.check_output(
            stage_grep_REGOUT_wns, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_wns = ans1_REGOUT_wns.rstrip("\n")
        if ans1_REGOUT_wns == "":
            ans1_REGOUT_wns = "NA"
    except:
        ans1_REGOUT_wns = "NA"

    try:
        ans1_REGOUT_tns = subprocess.check_output(
            stage_grep_REGOUT_tns, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_tns = ans1_REGOUT_tns.rstrip("\n")
        if ans1_REGOUT_tns == "":
            ans1_REGOUT_tns = "NA"
    except:
        ans1_REGOUT_tns = "NA"

    try:
        ans1_REGOUT_nve = subprocess.check_output(
            stage_grep_REGOUT_nve, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_nve = ans1_REGOUT_nve.rstrip("\n")
        if ans1_REGOUT_nve == "":
            ans1_REGOUT_nve = "NA"
    except:
        ans1_REGOUT_nve = "NA"
        
    try:
        ans1_wns_81v_125C = subprocess.check_output(
            stage_grep_wns_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        #print("THE aans1_wns_81v_125C",ans1_wns_81v_125C);
        ans1_wns_81v_125C = ans1_wns_81v_125C.rstrip("\n")
        if ans1_wns_81v_125C == "":
            ans1_wns_81v_125C = "NA"
    except:
        ans1_wns_81v_125C = "NA"

    try:
        ans1_tns_81v_125C = subprocess.check_output(
            stage_grep_tns_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_tns_81v_125C = ans1_tns_81v_125C.rstrip("\n")
        if ans1_tns_81v_125C == "":
            ans1_tns_81v_125C = "NA"
    except:
        ans1_tns_81v_125C = "NA"

    try:
        ans1_nve_81v_125C = subprocess.check_output(
            stage_grep_nve_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_nve_81v_125C = ans1_nve_81v_125C.rstrip("\n")
        if ans1_nve_81v_125C == "":
            ans1_nve_81v_125C = "NA"
    except:
        ans1_nve_81v_125C = "NA"

    try:
        ans1_REGIN_wns_81v_125C = subprocess.check_output(
            stage_grep_REGIN_wns_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_wns_81v_125C = ans1_REGIN_wns_81v_125C.rstrip("\n")
        if ans1_REGIN_wns_81v_125C == "":
            ans1_REGIN_wns_81v_125C = "NA"
    except:
        ans1_REGIN_wns_81v_125C = "NA"

    try:
        ans1_REGIN_tns_81v_125C = subprocess.check_output(
            stage_grep_REGIN_tns_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_tns_81v_125C = ans1_REGIN_tns_81v_125C.rstrip("\n")
        if ans1_REGIN_tns_81v_125C == "":
            ans1_REGIN_tns_81v_125C = "NA"
    except:
        ans1_REGIN_tns_81v_125C = "NA"

    try:
        ans1_REGIN_nve_81v_125C = subprocess.check_output(
            stage_grep_REGIN_nve_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_nve_81v_125C = ans1_REGIN_nve_81v_125C.rstrip("\n")
        if ans1_REGIN_nve_81v_125C == "":
            ans1_REGIN_nve_81v_125C = "NA"
    except:
        ans1_REGIN_nve_81v_125C = "NA"

    try:
        ans1_REGOUT_wns_81v_125C = subprocess.check_output(
            stage_grep_REGOUT_wns_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_wns_81v_125C = ans1_REGOUT_wns_81v_125C.rstrip("\n")
        if ans1_REGOUT_wns_81v_125C == "":
            ans1_REGOUT_wns_81v_125C = "NA"
    except:
        ans1_REGOUT_wns_81v_125C = "NA"

    try:
        ans1_REGOUT_tns_81v_125C = subprocess.check_output(
            stage_grep_REGOUT_tns_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_tns_81v_125C = ans1_REGOUT_tns_81v_125C.rstrip("\n")
        if ans1_REGOUT_tns_81v_125C == "":
            ans1_REGOUT_tns_81v_125C = "NA"
    except:
        ans1_REGOUT_tns_81v_125C = "NA"

    try:
        ans1_REGOUT_nve_81v_125C = subprocess.check_output(
            stage_grep_REGOUT_nve_81v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_nve_81v_125C = ans1_REGOUT_nve_81v_125C.rstrip("\n")
        if ans1_REGOUT_nve_81v_125C == "":
            ans1_REGOUT_nve_81v_125C = "NA"
    except:
        ans1_REGOUT_nve_81v_125C = "NA"
    
    try:
        ans1_wns_72v_125C = subprocess.check_output(
            stage_grep_wns_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_wns_72v_125C = ans1_wns_72v_125C.rstrip("\n")
        if ans1_wns_72v_125C == "":
            ans1_wns_72v_125C = "NA"
    except:
        ans1_wns_72v_125C = "NA"

    try:
        ans1_tns_72v_125C = subprocess.check_output(
            stage_grep_tns_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_tns_72v_125C = ans1_tns_72v_125C.rstrip("\n")
        if ans1_tns_72v_125C == "":
            ans1_tns_72v_125C = "NA"
    except:
        ans1_tns_72v_125C = "NA"

    try:
        ans1_nve_72v_125C = subprocess.check_output(
            stage_grep_nve_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_nve_72v_125C = ans1_nve_72v_125C.rstrip("\n")
        if ans1_nve_72v_125C == "":
            ans1_nve_72v_125C = "NA"
    except:
        ans1_nve_72v_125C = "NA"

    try:
        ans1_REGIN_wns_72v_125C = subprocess.check_output(
            stage_grep_REGIN_wns_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_wns_72v_125C = ans1_REGIN_wns_72v_125C.rstrip("\n")
        if ans1_REGIN_wns_72v_125C == "":
            ans1_REGIN_wns_72v_125C = "NA"
    except:
        ans1_REGIN_wns_72v_125C = "NA"

    try:
        ans1_REGIN_tns_72v_125C = subprocess.check_output(
            stage_grep_REGIN_tns_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_tns_72v_125C = ans1_REGIN_tns_72v_125C.rstrip("\n")
        if ans1_REGIN_tns_72v_125C == "":
            ans1_REGIN_tns_72v_125C = "NA"
    except:
        ans1_REGIN_tns_72v_125C = "NA"

    try:
        ans1_REGIN_nve_72v_125C = subprocess.check_output(
            stage_grep_REGIN_nve_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGIN_nve_72v_125C = ans1_REGIN_nve_72v_125C.rstrip("\n")
        if ans1_REGIN_nve_72v_125C == "":
            ans1_REGIN_nve_72v_125C = "NA"
    except:
        ans1_REGIN_nve_72v_125C = "NA"

    try:
        ans1_REGOUT_wns_72v_125C = subprocess.check_output(
            stage_grep_REGOUT_wns_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_wns_72v_125C = ans1_REGOUT_wns_72v_125C.rstrip("\n")
        if ans1_REGOUT_wns_72v_125C == "":
            ans1_REGOUT_wns_72v_125C = "NA"
    except:
        ans1_REGOUT_wns_72v_125C = "NA"

    try:
        ans1_REGOUT_tns_72v_125C = subprocess.check_output(
            stage_grep_REGOUT_tns_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_tns_72v_125C = ans1_REGOUT_tns_72v_125C.rstrip("\n")
        if ans1_REGOUT_tns_72v_125C == "":
            ans1_REGOUT_tns_72v_125C = "NA"
    except:
        ans1_REGOUT_tns_72v_125C = "NA"

    try:
        ans1_REGOUT_nve_72v_125C = subprocess.check_output(
            stage_grep_REGOUT_nve_72v_125C, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1_REGOUT_nve_72v_125C = ans1_REGOUT_nve_72v_125C.rstrip("\n")
        if ans1_REGOUT_nve_72v_125C == "":
            ans1_REGOUT_nve_72v_125C = "NA"
    except:
        ans1_REGOUT_nve_72v_125C = "NA"
    
    
    try:
        ans1 = subprocess.check_output(
            stage_grep_util, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans1 = ans1.rstrip("\n")
        if ans1 == "":
            ans1 = "NA"
    except:
        ans1 = "NA"

    try:
        ans2_h = subprocess.check_output(
            stage_grep_h, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_h = ans2_h.rstrip("\n")
        if ans2_h == "":
            ans2_h = "NA"
    except:
        ans2_h = "NA"

    try:
        ans2_v = subprocess.check_output(
            stage_grep_v, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_v = ans2_v.rstrip("\n")
        if ans2_v == "":
            ans2_v = "NA"
    except:
        ans2_v = "NA"

    try:
        ans2_std_count = subprocess.check_output(
            stage_grep_std_count, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_std_count = ans2_std_count.rstrip("\n")
        if ans2_std_count == "":
            ans2_std_count = "NA"
    except:
        ans2_std_count = "NA"

    try:
        ans2_std_area = subprocess.check_output(
            stage_grep_std_area, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_std_area = ans2_std_area.rstrip("\n")
        if ans2_std_area == "":
            ans2_std_area = "NA"
    except:
        ans2_std_area = "NA"

    try:
        ans2_AOB = subprocess.check_output(
            stage_grep_AOB, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_AOB = ans2_AOB.rstrip("\n")
        if ans2_AOB == "":
            ans2_AOB = "NA"
    except:
        ans2_AOB = "NA"

    try:
        ans2_both_dir = subprocess.check_output(
            stage_grep_both_dir, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_both_dir = ans2_both_dir.rstrip("\n")
        if ans2_both_dir == "":
            ans2_both_dir = "NA"
    except:
        ans2_both_dir = "NA"

    try:
        ans2_short = subprocess.check_output(
            stage_grep_Shorts, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_short = ans2_short.rstrip("\n")
        if ans2_short == "":
            ans2_short = "NA"
    except subprocess.CalledProcessError:
        ans2_short = "NA"

    try:
        ans2_DRC = subprocess.check_output(
            stage_grep_DRC, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_DRC = ans2_DRC.rstrip("\n")
        if ans2_DRC == "":
            ans2_DRC = "NA"

    except:
        ans2_DRC = "NA"

    try:
        ans2_bbox = subprocess.check_output(
            stage_grep_bbox, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_bbox = ans2_bbox.rstrip("\n")
        if ans2_bbox == "":
            ans2_bbox = "NA"

    except:
        ans2_bbox = "NA"
        
    try:
        ans2_lvt = subprocess.check_output(
            stage_grep_LVT, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_lvt = ans2_lvt.rstrip("\n")
        #print("UUUUUUUUUUUUUU", ans2_lvt)
        if ans2_lvt == "":
            ans2_lvt = "NA"

    except:
        ans2_lvt = "NA"
        
    try:
        ans2_svt = subprocess.check_output(
            stage_grep_SVT, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_svt = ans2_svt.rstrip("\n")
        if ans2_svt == "":
            ans2_svt = "NA"

    except:
        ans2_svt = "NA"
        
    try:
        ans2_ulvt = subprocess.check_output(
            stage_grep_ULVT, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_ulvt = ans2_ulvt.rstrip("\n")
        if ans2_ulvt == "":
            ans2_ulvt = "NA"

    except:
        ans2_ulvt = "NA"

    try:
        ans2_lvt16 = subprocess.check_output(
            stage_grep_LVT16, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_lvt16 = ans2_lvt16.rstrip("\n")
        #print("UUUUUUUUUUUUUU", ans2_lvt)
        if ans2_lvt16 == "":
            ans2_lvt16 = "NA"

    except:
        ans2_lvt16 = "NA"
        
    try:
        ans2_svt16 = subprocess.check_output(
            stage_grep_SVT16, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_svt16 = ans2_svt16.rstrip("\n")
        if ans2_svt16 == "":
            ans2_svt16 = "NA"

    except:
        ans2_svt16 = "NA"
        
    try:
        ans2_ulvt16 = subprocess.check_output(
            stage_grep_ULVT16, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_ulvt16 = ans2_ulvt16.rstrip("\n")
        if ans2_ulvt16 == "":
            ans2_ulvt16 = "NA"

    except:
        ans2_ulvt16 = "NA"

    try:
        ans2_lvt18 = subprocess.check_output(
            stage_grep_LVT18, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_lvt18 = ans2_lvt18.rstrip("\n")
        #print("UUUUUUUUUUUUUU", ans2_lvt)
        if ans2_lvt18 == "":
            ans2_lvt18 = "NA"

    except:
        ans2_lvt18 = "NA"
        
    try:
        ans2_svt18 = subprocess.check_output(
            stage_grep_SVT18, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_svt18 = ans2_svt18.rstrip("\n")
        if ans2_svt18 == "":
            ans2_svt18 = "NA"

    except:
        ans2_svt18 = "NA"
        
    try:
        ans2_ulvt18 = subprocess.check_output(
            stage_grep_ULVT18, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_ulvt18 = ans2_ulvt18.rstrip("\n")
        if ans2_ulvt18 == "":
            ans2_ulvt18 = "NA"

    except:
        ans2_ulvt18 = "NA"

    try:
        ans2_lvt20 = subprocess.check_output(
            stage_grep_LVT20, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_lvt20 = ans2_lvt20.rstrip("\n")
        #print("UUUUUUUUUUUUUU", ans2_lvt)
        if ans2_lvt20 == "":
            ans2_lvt20 = "NA"

    except:
        ans2_lvt20 = "NA"
        
    try:
        ans2_svt20 = subprocess.check_output(
            stage_grep_SVT20, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_svt20 = ans2_svt20.rstrip("\n")
        if ans2_svt20 == "":
            ans2_svt20 = "NA"

    except:
        ans2_svt20 = "NA"
        
    try:
        ans2_ulvt20 = subprocess.check_output(
            stage_grep_ULVT20, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_ulvt20 = ans2_ulvt20.rstrip("\n")
        if ans2_ulvt20 == "":
            ans2_ulvt20 = "NA"

    except:
        ans2_ulvt20 = "NA"

    try:
        ans2_lvt24 = subprocess.check_output(
            stage_grep_LVT24, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_lvt24 = ans2_lvt24.rstrip("\n")
        #print("UUUUUUUUUUUUUU", ans2_lvt24)
        if ans2_lvt24 == "":
            ans2_lvt24 = "NA"

    except:
        ans2_lvt24 = "NA"
        
    try:
        ans2_svt24 = subprocess.check_output(
            stage_grep_SVT24, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_svt24 = ans2_svt24.rstrip("\n")
        if ans2_svt24 == "":
            ans2_svt24 = "NA"

    except:
        ans2_svt24 = "NA"
        
    try:
        ans2_ulvt24 = subprocess.check_output(
            stage_grep_ULVT24, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_ulvt24 = ans2_ulvt24.rstrip("\n")
        if ans2_ulvt24 == "":
            ans2_ulvt24 = "NA"

    except:
        ans2_ulvt24 = "NA"

    try:
        ans2_ccd = subprocess.check_output(
            stage_grep_ccd, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_ccd = ans2_ccd.rstrip("\n")
        if ans2_ccd == "0":
            ans2_ccd = "NO"
        else:
            ans2_ccd = "YES"

    except subprocess.CalledProcessError:
        ans2_ccd = "No"

    #print("Ans1 is", ans1)
    #print("H Ans is", ans2_h)
    #print("V Ans is", ans2_v)

    #print("std_count ans", ans2_std_count)
    #print("std_cell area", ans2_std_area)
    #print("AOB bufer # :", ans2_AOB)
    
    #print("the ans_lvt is",ans2_lvt);

    stage_util.append(ans1)
    stage_h.append(ans2_h)
    stage_v.append(ans2_v)
    stage_std_count.append(ans2_std_count)
    stage_std_cell_area.append(ans2_std_area)
    stage_AOB.append(ans2_AOB)
    stage_shorts.append(ans2_short)
    stage_both_dir.append(ans2_both_dir)
    stage_DRC.append(ans2_DRC)
    stage_overall.append(f"{ans2_both_dir} |( {ans2_DRC} / {ans2_short} )")
    stage_ccd.append(ans2_ccd)
    stage_bbox.append(ans2_bbox)
    stage_REG2REG.append(f"{ans1_wns} | {ans1_tns} | {ans1_nve}")
    stage_REG2IN.append(f"{ans1_REGIN_wns} | {ans1_REGIN_tns} | {ans1_REGIN_nve}")
    stage_REG2OUT.append(f"{ans1_REGOUT_wns} | {ans1_REGOUT_tns} | {ans1_REGOUT_nve}")
    
    stage_REG2REG_72v_125C.append(f"{ans1_wns_72v_125C} | {ans1_tns_72v_125C} | {ans1_nve_72v_125C}")
    stage_REG2IN_72v_125C.append(f"{ans1_REGIN_wns_72v_125C} | {ans1_REGIN_tns_72v_125C} | {ans1_REGIN_nve_72v_125C}")
    stage_REG2OUT_72v_125C.append(f"{ans1_REGOUT_wns_72v_125C} | {ans1_REGOUT_tns_72v_125C} | {ans1_REGOUT_nve_72v_125C}")
    
    stage_REG2REG_81v_125C.append(f"{ans1_wns_81v_125C} | {ans1_tns_81v_125C} | {ans1_nve_81v_125C}")
    stage_REG2IN_81v_125C.append(f"{ans1_REGIN_wns_81v_125C} | {ans1_REGIN_tns_81v_125C} | {ans1_REGIN_nve_81v_125C}")
    stage_REG2OUT_81v_125C.append(f"{ans1_REGOUT_wns_81v_125C} | {ans1_REGOUT_tns_81v_125C} | {ans1_REGOUT_nve_81v_125C}")
    
    if ans2_svt != "NA":
            ans2_svt = float(ans2_svt)
            ans2_svt = int(ans2_svt * 10 ** 2) / 10 ** 2
            ans2_svt = str(ans2_svt);
    if ans2_lvt != "NA":
            ans2_lvt = float(ans2_lvt)
            ans2_lvt = int(ans2_lvt * 10 ** 2) / 10 ** 2
            ans2_lvt = str(ans2_lvt);
    if ans2_ulvt != "NA":
            ans2_ulvt = float(ans2_ulvt)
            ans2_ulvt = int(ans2_ulvt * 10 ** 2) / 10 ** 2
            ans2_ulvt = str(ans2_ulvt);
    
    #if ans2_svt16 != "NA":
            #ans2_svt16 = float(ans2_svt16)
            #ans2_svt16 = int(ans2_svt16 * 10 ** 2) / 10 ** 2
            #ans2_svt16 = str(ans2_svt16);
    #if ans2_lvt16 != "NA":
            #ans2_lvt16 = float(ans2_lvt16)
            #ans2_lvt16 = int(ans2_lvt16 * 10 ** 2) / 10 ** 2
            #ans2_lvt16 = str(ans2_lvt16);
    #if ans2_ulvt16 != "NA":
            #ans2_ulvt16 = float(ans2_ulvt16)
            #ans2_ulvt16 = int(ans2_ulvt16 * 10 ** 2) / 10 ** 2
            #ans2_ulvt16 = str(ans2_ulvt16);

    #if ans2_svt18 != "NA":
            #ans2_svt18 = float(ans2_svt18)
            #ans2_svt18 = int(ans2_svt18 * 10 ** 2) / 10 ** 2
            #ans2_svt18 = str(ans2_svt18);
    #if ans2_lvt18 != "NA":
            #ans2_lvt18 = float(ans2_lvt18)
            #ans2_lvt18 = int(ans2_lvt18 * 10 ** 2) / 10 ** 2
            #ans2_lvt18 = str(ans2_lvt18);
    #if ans2_ulvt18 != "NA":
            #ans2_ulvt18 = float(ans2_ulvt18)
            #ans2_ulvt18 = int(ans2_ulvt18 * 10 ** 2) / 10 ** 2
            #ans2_ulvt18 = str(ans2_ulvt18);

    #if ans2_svt20 != "NA":
            #ans2_svt20 = float(ans2_svt20)
            #ans2_svt20 = int(ans2_svt20 * 10 ** 2) / 10 ** 2
            #ans2_svt20 = str(ans2_svt20);
    #if ans2_lvt20 != "NA":
            #ans2_lvt20 = float(ans2_lvt20)
            #ans2_lvt20 = int(ans2_lvt20 * 10 ** 2) / 10 ** 2
            #ans2_lvt20 = str(ans2_lvt20);
    #if ans2_ulvt20 != "NA":
            #ans2_ulvt20 = float(ans2_ulvt20)
            #ans2_ulvt20 = int(ans2_ulvt20 * 10 ** 2) / 10 ** 2
            #ans2_ulvt20 = str(ans2_ulvt20);

    #if ans2_svt24 != "NA":
            #ans2_svt24 = float(ans2_svt24)
            #ans2_svt24 = int(ans2_svt24 * 10 ** 2) / 10 ** 2
            #ans2_svt24 = str(ans2_svt24);
    #if ans2_lvt24 != "NA":
            #ans2_lvt24 = float(ans2_lvt24)
            #ans2_lvt24 = int(ans2_lvt24 * 10 ** 2) / 10 ** 2
            #ans2_lvt24 = str(ans2_lvt24);
    #if ans2_ulvt24 != "NA":
            #ans2_ulvt24 = float(ans2_ulvt24)
            #ans2_ulvt24 = int(ans2_ulvt24 * 10 ** 2) / 10 ** 2
            #ans2_ulvt24 = str(ans2_ulvt24);   

    stage_LVT.append(ans2_lvt);
    stage_SVT.append(ans2_svt)
    stage_ULVT.append(ans2_ulvt)

    stage_LVT24.append(ans2_lvt24);
    stage_SVT24.append(ans2_svt24)
    stage_ULVT24.append(ans2_ulvt24)

    stage_LVT16.append(ans2_lvt16);
    stage_SVT16.append(ans2_svt16)
    stage_ULVT16.append(ans2_ulvt16)   

    stage_LVT18.append(ans2_lvt18);
    stage_SVT18.append(ans2_svt18)
    stage_ULVT18.append(ans2_ulvt18)   

    stage_LVT20.append(ans2_lvt20);
    stage_SVT20.append(ans2_svt20)
    stage_ULVT20.append(ans2_ulvt20)      

#print("The reg 2reg is", stage_REG2REG)

#print("The D_name is ", d_name)
#print(stage_util)

#print("the Stage H is ", stage_h)
#print("the stage V is ", stage_v)


#print("The stnd coun is ", stage_std_count)
#print("The stnd cell area", stage_std_cell_area)
#print("the AOB is :", stage_AOB)
#print("the Shorts is ", stage_shorts)
#print("the overall/ DRC", stage_overall)

#print("The stage LVT is", stage_LVT);
print("\n\n")


#print("The log pat is ", log)

stage_Error = []
stage_Warning = []
stage_version = []
stage_runtime = []
stage_mem_usage = []
for i in range(0, len(stage_name) - 1):
    try:
        stage_grep_Error = f'grep -c "^Error" {log}/logs/{d_name[i]}/{log_ext[i]}.log'
        # grep -c "^Error" 35_pnr/logs/{d_name[i]}/icc2_*log
        #print(stage_grep_Error)
        ans2_Error = subprocess.check_output(
            stage_grep_Error, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_Error = ans2_Error.rstrip("\n")
        # ans2_Error= ans2_Error.lstrip(":");
        if ans2_Error == "":
            ans2_Error = "NA"
    except:
        ans2_Error = "NA"

    try:
        stage_grep_Warning = (
            f'grep -c "^[Ww]arning" {log}/logs/{d_name[i]}/{log_ext[i]}.log'
        )
        # grep -c "^Error" 35_pnr/logs/{d_name[i]}/icc2_*log
        #print(stage_grep_Warning)
        ans2_Warning = subprocess.check_output(
            stage_grep_Warning, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_Warning = ans2_Warning.rstrip("\n")
        # ans2_Error= ans2_Error.lstrip(":");
        if ans2_Warning == "":
            ans2_Warning = "NA"
    except:
        ans2_Warning = "NA"

    try:
        stage_grep_version = f'grep "Version" {log}/logs/{d_name[i]}/{log_ext[i]}.log | head -1 | sed -r "s/for.*//g" | sed -e "s/^[ ]*//"'
        # grep "Version" 35_pnr/logs/720_icc2_route_atomic/icc2_route_atomic.log | head -1 | sed -r "s/for.*//g" | sed -e "s/^[ ]*//"
        #print(stage_grep_version)
        ans2_version = subprocess.check_output(
            stage_grep_version, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        #print(ans2_version)
        ans2_version = ans2_version.rstrip("\n")
        # ans2_Error= ans2_Error.lstrip(":");
        if ans2_version == "":
            ans2_version = "NA"
    except:
        ans2_version = "NA"

    try:
        stage_grep_runtime = f'tail {log}/logs/{d_name[i]}/{log_ext[i]}.log | grep "Elapsed" | grep -oE "[0-9]+\.[0-9]+"'  ##grep "Runtime" 35_pnr/rpts/*/QOR_imx95_wakeupmix_wrapper.csv  | grep -oE "[0-9]+\.[0-9]+"
        ans2_runtime = subprocess.check_output(
            stage_grep_runtime, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_runtime = ans2_runtime.rstrip("\n")
        if ans2_runtime == "":
            ans2_runtime = "NA"
    except:
        ans2_runtime = "NA"

    try:
        stage_grep_mem_usage = f'tail {log}/logs/{d_name[i]}/{log_ext[i]}.log | grep "CPU usage" | grep -oE "[0-9]+\.[0-9]+"'  ##grep "Runtime" 35_pnr/rpts/*/QOR_imx95_wakeupmix_wrapper.csv  | grep -oE "[0-9]+\.[0-9]+"
        ans2_mem = subprocess.check_output(
            stage_grep_mem_usage, stdin=subprocess.PIPE, shell=True
        ).decode("utf-8")
        ans2_mem = ans2_mem.rstrip("\n")
        if ans2_mem == "":
            ans2_mem = "NA"
    except:
        ans2_mem = "NA"
    stage_Error.append(ans2_Error)
    stage_Warning.append(ans2_Warning)
    stage_version.append(ans2_version)
    stage_runtime.append(ans2_runtime)
    stage_mem_usage.append(ans2_mem)

#print("The stage Error is", stage_Error)
#print("the Warning count is", stage_Warning)
#print("the stage_version", stage_version)
#print("The runtime_version", stage_runtime)
#print("The mem usage is ", stage_mem_usage)
work_obj = Workbook()

sheet = work_obj.active

sheet.append(["RUN DIR", curr_dir]);
sheet.merge_cells('B1:O1');
sheet.title = "Task_pnr"
Des_name = {
    "title": "Design_Name",
    "Value": Des_Name_li,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

HM = {
    "title": "Tag Name",
    "Value": design_name_li,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

stage_name = {
    "title": "stage_name",
    "Value": d_name,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Standard_cell_Count = {
    "title": "Standard_cell_Count",
    "Value": stage_std_count,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
# Congestion["Width"] = len(Congestion["title"]) +1;

Standard_cell_Area = {
    "title": "Standard_cell_Area",
    "Value": stage_std_cell_area,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

AOB_buffer = {
    "title": "AOB Buffers",
    "Value": stage_AOB,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

ccd = {
    "title": "CCD Enabled",
    "Value": stage_ccd,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

bbox = {
    "title": "bbox (Chip area)",
    "Value": stage_bbox,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Utilization = {
    "title": "Utilization_value",
    "Value": stage_util,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [0.657, 0.711],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
# Utilization["Width"] =len(Utilization["title"]) +1;

Error_Count = {
    "title": "Error_Count",
    "Value": stage_Error,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Warning_Count = {
    "title": "Warning_Count",
    "Value": stage_Warning,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

ICC_version = {
    "title": "ICC_version",
    "Value": stage_version,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Runtime = {
    "title": "Runtime in hrs",
    "Value": stage_runtime,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
Cpu_usage = {
    "title": "CPU Usage in hrs",
    "Value": stage_mem_usage,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}


# Congestion_Image["Width"] =len(Congestion_Image["title"]) +1;


H = {
    "title": "H_ Routing %",
    "Value": stage_h,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

V = {
    "title": "V_Routing %",
    "Value": stage_v,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Overall = {
    "title": "Overall/(DRC/Short)",
    "Value": stage_overall,
    "Width": "60",
    "Row": [],
    "col": "F",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}


Reg2reg_72v_40C = {
    "title": "REG2REG",
    "Value": stage_REG2REG,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2in_72v_40C = {
    "title": "REGIN",
    "Value": stage_REG2IN,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2out_72v_40C = {
    "title": "REGOUT",
    "Value": stage_REG2OUT,
    "Width": "60",
    "Row": [],
    "col": "F",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "align": Alignment(horizontal="center", vertical="center"),
    "subtitle": [],
}
    
Reg2reg_81v_125C = {
    "title": "REG2REG",
    "Value": stage_REG2REG_81v_125C,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2in_81v_125C = {
    "title": "REGIN",
    "Value": stage_REG2IN_81v_125C,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2out_81v_125C = {
    "title": "REGOUT",
    "Value": stage_REG2OUT_81v_125C,
    "Width": "60",
    "Row": [],
    "col": "F",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "align": Alignment(horizontal="center", vertical="center"),
    "subtitle": [],
}
    
Reg2reg_72v_125C = {
    "title": "REG2REG",
    "Value": stage_REG2REG_72v_125C,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2in_72v_125C = {
    "title": "REGIN",
    "Value": stage_REG2IN_72v_125C,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2out_72v_125C = {
    "title": "REGOUT",
    "Value": stage_REG2OUT_72v_125C,
    "Width": "60",
    "Row": [],
    "col": "F",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "align": Alignment(horizontal="center", vertical="center"),
    "subtitle": [],
}

Overall2 = {
    "title": "Overall/(DRC/Short)",
    "Value": [],
    "Width": "60",
    "Row": [],
    "col": "F",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Congestion_topic = {
    "title": "Congestion_topic",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [H, V, Overall],
}

Timing_scenario_72v_40C = {
    "title": "TIMING SCENARIO(FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T)",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [Reg2reg_72v_40C, Reg2in_72v_40C, Reg2out_72v_40C],
}

Timing_scenario_72v_125C = {
    "title": "TIMING SCENARIO(FUNC_NM.SSGNP_0p72v_125C.rcworst_CCworst_T)",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [Reg2reg_72v_125C, Reg2in_72v_125C, Reg2out_72v_125C],
}
    
Timing_scenario_81v_125C = {
    "title": "TIMING SCENARIO(FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T)",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [Reg2reg_81v_125C, Reg2in_81v_125C, Reg2out_81v_125C],
}

# merge_dic = {
# "Congestion_topic" : len(Congestion_topic["subtitle"]),
# "Congestion_topic1": len(Congestion_topic1["subtitle"])
# };

LVT = {
    "title": "LVT Area %(Overall)",
    "Value": stage_LVT,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

SVT= {
    "title": "SVT Area %(Overall)",
    "Value": stage_SVT,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

ULVT = {
    "title": "ULVT Area % (Overall)",
    "Value": stage_ULVT,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

LVT16 = {
    "title": "LVT_16 Area %",
    "Value": stage_LVT16,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

SVT16= {
    "title": "SVT_16 Area %",
    "Value": stage_SVT16,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

ULVT16 = {
    "title": "ULVT_16 Area %",
    "Value": stage_ULVT16,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

LVT18 = {
    "title": "LVT_18 Area %",
    "Value": stage_LVT18,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

SVT18= {
    "title": "SVT_18 Area %",
    "Value": stage_SVT18,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

ULVT18 = {
    "title": "ULVT_18 Area %",
    "Value": stage_ULVT18,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

LVT20 = {
    "title": "LVT_20 Area %",
    "Value": stage_LVT20,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

SVT20= {
    "title": "SVT_20 Area %",
    "Value": stage_SVT20,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

ULVT20 = {
    "title": "ULVT_20 Area %",
    "Value": stage_ULVT20,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
LVT24 = {
    "title": "LVT_24 Area %",
    "Value": stage_LVT24,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

SVT24= {
    "title": "SVT_24 Area %",
    "Value": stage_SVT24,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

ULVT24 = {
    "title": "ULVT_24 Area %",
    "Value": stage_ULVT24,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

All_LVT = {
    "title": "LVT",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [LVT, LVT16, LVT18, LVT20, LVT24],
}
All_SVT = {
    "title": "SVT",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [SVT, SVT16, SVT18, SVT20, SVT24],
}
All_ULVT = {
    "title": "ULVT",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [ULVT, ULVT16, ULVT18, ULVT20, ULVT24],
}
    


H3 = {
    "title": "H",
    "Value": [],
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

V3 = {
    "title": "V",
    "Value": [],
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Overall3 = {
    "title": "Overall/(DRC/Short)",
    "Value": [],
    "Width": "60",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

X3 = {
    "title": "X3",
    "Value": [],
    "Width": "60",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Testing1 = {
    "title": "Testing1",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [H3, V3, Overall3, X3],
}


av = [
    Des_name,
    HM,
    stage_name,
    Utilization,
    Standard_cell_Count,
    Standard_cell_Area,
    AOB_buffer,
    ccd,
    bbox,
    Congestion_topic,
    Error_Count,
    Warning_Count,
    ICC_version,
    Runtime,
    Cpu_usage,
    Timing_scenario_72v_40C,
    Timing_scenario_72v_125C,
    Timing_scenario_81v_125C,
    All_LVT,
    All_SVT,
    All_ULVT,
    
]

###--------------------------------------------------------------------------------------------
#for Synthesis

sheet_syn = work_obj.create_sheet("Task_syn");

Des_name_syn = {
    "title": "Design_Name",
    "Value": Des_Name_li_syn,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

HM_syn = {
    "title": "Tag Name",
    "Value": design_name_li_syn,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

stage_name_syn = {
    "title": "stage_name",
    "Value": d_name_syn,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
    
One_bit_eq_gated_regs = {
    "title": "1_bit eq gated regs %",
    "Value": stage_gate_reg,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2reg_72v_40C_syn = {
    "title": "REG2REG",
    "Value": stage_REG2REG_syn,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2in_72v_40C_syn = {
    "title": "REGIN",
    "Value": stage_REG2IN_syn,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2out_72v_40C_syn= {
    "title": "REGOUT",
    "Value": stage_REG2OUT_syn,
    "Width": "60",
    "Row": [],
    "col": "F",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "align": Alignment(horizontal="center", vertical="center"),
    "subtitle": [],
}
    
Timing_scenario_72v_40C_syn = {
    "title": "TIMING SCENARIO (FUNC_NM.SSGNP_0p72v_m40C.cworst_CCworst_T)",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [Reg2reg_72v_40C_syn, Reg2in_72v_40C_syn, Reg2out_72v_40C_syn],
}
    
    
Reg2reg_81v_125C_syn = {
    "title": "REG2REG",
    "Value": stage_REG2REG_81v_125C_syn,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2in_81v_125C_syn = {
    "title": "REGIN",
    "Value": stage_REG2IN_81v_125C_syn,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2out_81v_125C_syn = {
    "title": "REGOUT",
    "Value": stage_REG2OUT_81v_125C_syn,
    "Width": "60",
    "Row": [],
    "col": "F",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "align": Alignment(horizontal="center", vertical="center"),
    "subtitle": [],
}
    
Reg2reg_72v_125C_syn = {
    "title": "REG2REG",
    "Value": stage_REG2REG_72v_125C_syn,
    "Width": "20",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2in_72v_125C_syn = {
    "title": "REGIN",
    "Value": stage_REG2IN_72v_125C_syn,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Reg2out_72v_125C_syn = {
    "title": "REGOUT",
    "Value": stage_REG2OUT_72v_125C_syn,
    "Width": "60",
    "Row": [],
    "col": "F",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "align": Alignment(horizontal="center", vertical="center"),
    "subtitle": [],
}




Timing_scenario_72v_125C_syn = {
    "title": "TIMING SCENARIO(FUNC_NM.SSGNP_0p72v_125C.cworst_CCworst_T)",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [Reg2reg_72v_125C_syn, Reg2in_72v_125C_syn, Reg2out_72v_125C_syn],
}
    
Timing_scenario_81v_125C_syn = {
    "title": "TIMING SCENARIO(FUNC_OD.SSGNP_0p81v_125C.rcworst_CCworst_T)",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [Reg2reg_81v_125C_syn, Reg2in_81v_125C_syn, Reg2out_81v_125C_syn],
}

Bit_1 = {
    "title": "1-Bit %",
    "Value": stage_1_bit_syn,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
    
Bit_2 = {
    "title": "2-Bit %",
    "Value": stage_2_bit_syn,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Bit_4 = {
    "title": "4-Bit %",
    "Value": stage_4_bit_syn,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
    
Bit_8 = {
    "title": "8-Bit %",
    "Value": stage_8_bit_syn,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
    
Total_Bit = {
    "title": "Total-Bit Count",
    "Value": stage_Total_bit,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}


multibit_flop_count = {
    "title": "%_of Total flop count(Using 1-bit equivalent numbers)",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "D",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [Bit_1,Bit_2, Bit_4, Bit_8, Total_Bit],
} 
    
DC_version_syn = {
    "title": "DC_version",
    "Value": stage_version_syn,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
    
dft_tracing = {
    "title": "DFT_Tracing",
    "Value": stage_dft_trace,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
FE_bom = {
    "title": "FE BOM",
    "Value": stage_FE_Bom_syn,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Port_BUFFER = {
    "title": "Port Buffer",
    "Value": stage_port_buf,
    "Width": "20",
    "Row": [],
    "col": "E",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}


av_syn = [
  Des_name_syn,
  stage_name_syn,
  One_bit_eq_gated_regs,
  Timing_scenario_72v_40C_syn,
  Timing_scenario_72v_125C_syn,
  Timing_scenario_81v_125C_syn,
  DC_version_syn,
  multibit_flop_count,
  dft_tracing,
  FE_bom,
  Port_BUFFER,
  ]

for i in range(0, len(av_syn)):
    av_syn[i]["col"] = chr(65 + i)
    av_syn[i]["Width"] = len(av_syn[i]["title"]) + 2
    # Changing the width of the column based on the title length


row_li = [1, 2]

for i in range(3, len(d_name_syn) + 3):
    row_li.append(i)
# print("The row li is :")
# print(row_li)

for C in av_syn:
    if C["subtitle"] == []:
        C["Row"] = row_li
    else:
        C["Row"] = row_li
        for c in C["subtitle"]:
            c["Row"] = row_li


# Code for Assigning col value to each dictionary
is_array = True
i = 0
k = 0
while is_array:
    if k == len(av_syn) - 1:
        is_array = False

    if av_syn[k]["subtitle"] == []:
        av_syn[k]["col"] = get_column_letter(1 + i)
        i = i + 1
    else:
        av_syn[k]["col"] = get_column_letter(1 + i)

        j = 0
        for sub_title in av_syn[k]["subtitle"]:
            sub_title["col"] = get_column_letter(1 + i + j)
            j = j + 1
        i = i + j
    # temp= av_syn[k]["title"]
    k = k + 1


# code for merging the desired cells
for dic in av_syn:
    if dic["subtitle"] == []:
        ans = f'{dic["col"]}{dic["Row"][0]}:{dic["col"]}{dic["Row"][1]}'
        sheet_syn.merge_cells(ans)
    else:
        ans = f'{dic["subtitle"][0]["col"]}1:{dic["subtitle"][-1]["col"]}1'
        # print(ans, type(ans), end= " ")
        sheet_syn.merge_cells(ans)

#print("YYYY")

# code for adding Title , value,font, fill, align,
for title in av_syn:
    if title["subtitle"] == []:
        tit = f'{title["col"]}{title["Row"][0]}'
        # print("the title is")
        # print(title["col"], title["Row"][0]);
        sheet_syn[tit] = title["title"]

        sheet_syn[tit].font = title["font"]
        sheet_syn[tit].fill = title["fill"]
        sheet_syn[tit].alignment = title["align"]
        sheet_syn[tit].border = title["border"]

        # value = f'{title["col"]}{title["Row"][-1]}'
        # sheet_syn[value] = title["Value"];
    else:
        #print(title["col"], title["Row"][0])
        tit = f'{title["col"]}{title["Row"][0]}'
        # print("the SUB_title is")
        # print(title["col"], title["Row"][0])
        sheet_syn[tit] = title["title"]
        sheet_syn[tit].font = title["font"]
        sheet_syn[tit].fill = title["fill"]
        sheet_syn[tit].alignment = title["align"]
        sheet_syn[tit].border = title["border"]
        for sub in title["subtitle"]:
            tit = f'{sub["col"]}{sub["Row"][1]}'
            sheet_syn[tit] = sub["title"]

            sheet_syn[tit].font = sub["font"]
            sheet_syn[tit].fill = sub["fill"]
            sheet_syn[tit].alignment = sub["align"]
            sheet_syn[tit].border = sub["border"]
            # value = f'{sub["col"]}{sub["Row"][-1]}'
            # sheet_syn[value] = sub["Value"];

# inserting values

for i in av_syn:
    if i["Value"] == []:
        if i["subtitle"] == []:
            continue
        else:
            for j in i["subtitle"]:
                if j["Value"] == []:
                    continue
                else:
                    for k in range(0, len(j["Value"])):
                        A3 = f'{j["col"]}{j["Row"][k+2]}'
                        sheet_syn[A3].value = j["Value"][k]
    else:
        for j in range(0, len(i["Value"])):
            A3 = f'{i["col"]}{i["Row"][j+2]}'
            sheet_syn[A3].value = i["Value"][j]
            # print(i["Value"]);


#code for updating the cell width value
def call_csv(last_row_value):
    if is_csv == True:
        # Csv_title =[];\
        c = last_row_value
        for csv_file in csv_syn_li:
            if "VIEW" in csv_file:
                continue
            if csv_file == "":
                continue
            #print("1")

            Csv_title = list(csv_file.split("/"))[2]
            gap = 5
            c += gap
            row_c = c;
            r_val = f"A{c}"
            sheet_syn[r_val] = Csv_title
            csv_f = f"{curr_dir}/{csv_file}"
            with open(csv_f, mode="r") as file_obj:
                read_csv = csv.reader(file_obj)
                #print("The reader is", read_csv)
                count_r = 0
                for row_val in read_csv:
                    row_c = row_c +1;
                    if count_r == 0:
                        #l = f"A{c}:D{c}"
                        D = chr(65 + len(row_val) -1)
                        #print("The D is :",D)
                        l = f"A{c}:D{c}"
                        sheet_syn.merge_cells(l)
                        sheet_syn[r_val].font = title["font"]
                        sheet_syn[r_val].fill = title["fill"]
                        sheet_syn[r_val].alignment = title["align"]
                        sheet_syn[r_val].border = title["border"]
                        count_r = 1
                    sheet_syn.append(row_val)
                    
                    if (row_val != [] and row_val[-1] == "PASS"):
                        r_c = get_column_letter(len(row_val))
                        #print(r_c)
                        #print(type(r_c));
                        sheet_syn[f'{r_c}{row_c}'].fill = PatternFill(fill_type="solid",start_color='0000FF00');  #green
                    elif (row_val != [] and row_val[-1] == "FAIL"):
                        r_c = get_column_letter(len(row_val))
                        sheet_syn[f'{r_c}{row_c}'].fill = PatternFill(fill_type="solid",start_color='0FF00000');
                    c = c + 1
       
    else:
        c = last_row_value;
        gap = 5
        c += gap
        r_val = f"A{c}"
        sheet_syn[r_val] = " synth_check.csv file is NOT Present";
        l = f"A{c}:D{c}"
        sheet_syn.merge_cells(l)
        sheet_syn[r_val].font = Des_name_syn["font"];
        sheet_syn[r_val].fill = title["fill"];
        sheet_syn[r_val].alignment = title["align"]





call_csv(row_li[-1]);


for i in av_syn:
    if i["subtitle"] != []:
        for k in i["subtitle"]:
            max_len = 0
            for l in k["Value"]:
                if max_len < len(l):
                    max_len = len(l)
            if max_len < len(k["title"]):
                max_len = len(k["title"])
            sheet_syn.column_dimensions[f'{k["col"]}'].width = max_len+2;
        continue
    else:
        max_len = 0
        for j in i["Value"]:
            if max_len < len(j):
                max_len = len(j)
        if max_len < len(i["title"]):
            max_len = len(i["title"])
        sheet_syn.column_dimensions[f'{i["col"]}'].width = max_len+2;
      
# print("The max row value is :",sheet_syn.max_row)
# print("The maximun column value is:",sheet_syn.max_column)




###------------------------------------------------------------------------------------------------

#for i in range(0, len(av)):
    #av[i]["col"] = chr(65 + i)
    #av[i]["Width"] = len(av[i]["title"]) + 2
    ## Changing the width of the column based on the title length


row_li = [3, 4]

for i in range(row_li[-1]+1, len(d_name) + row_li[-1]+1):
    row_li.append(i)
# print("The row li is :")
# print(row_li)

for C in av:
    if C["subtitle"] == []:
        C["Row"] = row_li
    else:
        C["Row"] = row_li
        for c in C["subtitle"]:
            c["Row"] = row_li


# Code for Assigning col value to each dictionary
is_array = True
i = 0
k = 0
while is_array:
    if k == len(av) - 1:
        is_array = False

    if av[k]["subtitle"] == []:
        av[k]["col"] = get_column_letter(1 + i)
        i = i + 1
    else:
        av[k]["col"] = get_column_letter(1 + i)

        j = 0
        for sub_title in av[k]["subtitle"]:
            sub_title["col"] = get_column_letter(1 + i + j)
            j = j + 1
        i = i + j
    # temp= av[k]["title"]
    k = k + 1


# code for merging the desired cells
for dic in av:
    if dic["subtitle"] == []:
        ans = f'{dic["col"]}{dic["Row"][0]}:{dic["col"]}{dic["Row"][1]}'
        sheet.merge_cells(ans)
        #sheet.merge_cells(start_row= dic["Row"][0], start_column=int(dic["col"]), end_row=dic["Row"][1], end_column= int(dic["col"]))
    else:
        ans = f'{dic["subtitle"][0]["col"]}{dic["Row"][0]}:{dic["subtitle"][-1]["col"]}{dic["Row"][0]}'
        # print(ans, type(ans), end= " ")
        sheet.merge_cells(ans)
        #sheet.merge_cells(start_row= 1, start_column=int(dic["subtitle"][0]["col"]), end_row=1, end_column= int(dic["subtitle"][-1]["col"]))

#print("YYYY")

# code for adding Title , value,font, fill, align,
for title in av:
    if title["subtitle"] == []:
        tit = f'{title["col"]}{title["Row"][0]}'
        # print("the title is")
        # print(title["col"], title["Row"][0]);
        sheet[tit] = title["title"]

        sheet[tit].font = title["font"]
        sheet[tit].fill = title["fill"]
        sheet[tit].alignment = title["align"]
        sheet[tit].border = title["border"]

        # value = f'{title["col"]}{title["Row"][-1]}'
        # sheet[value] = title["Value"];
    else:
        #print(title["col"], title["Row"][0])
        tit = f'{title["col"]}{title["Row"][0]}'
        #print("the SUB_title is")
        #print(title["col"], title["Row"][0])
        sheet[tit] = title["title"]
        sheet[tit].font = title["font"]
        sheet[tit].fill = title["fill"]
        sheet[tit].alignment = title["align"]
        sheet[tit].border = title["border"]
        for sub in title["subtitle"]:
            tit = f'{sub["col"]}{sub["Row"][1]}'
            sheet[tit] = sub["title"]

            sheet[tit].font = sub["font"]
            sheet[tit].fill = sub["fill"]
            sheet[tit].alignment = sub["align"]
            sheet[tit].border = sub["border"]
            # value = f'{sub["col"]}{sub["Row"][-1]}'
            # sheet[value] = sub["Value"];

# inserting values

for i in av:
    if i["Value"] == []:
        if i["subtitle"] == []:
            continue
        else:
            for j in i["subtitle"]:
                if j["Value"] == []:
                    continue
                else:
                    for k in range(0, len(j["Value"])):
                        A3 = f'{j["col"]}{j["Row"][k+2]}'
                        sheet[A3].value = j["Value"][k]
                        if (j["Threshold"]  != []):
                            thres = float(j["Value"][k]);
                            if (thres < j["Threshold"][0]):
                                sheet[A3].font = Font(color= "00000000");
                                sheet[A3].fill =PatternFill(fill_type="solid",start_color='0000FF00') #Green
                            elif(thres >= j["Threshold"][0] and thres < j["Threshold"][1]):
                                sheet[A3].font = Font(color= "00000000");
                                sheet[A3].fill =PatternFill(fill_type="solid",start_color='00FFFF99'); #yellow
                            elif (thres >= j["Threshold"][1]):
                                sheet[A3].font = Font(color= "00000000");
                                sheet[A3].fill =PatternFill(fill_type="solid",start_color='00FF0000'); #RED
                        
    else:
        for j in range(0, len(i["Value"])):
            A3 = f'{i["col"]}{i["Row"][j+2]}'
            sheet[A3].value = i["Value"][j]
            if (i["Threshold"]  != []):
                thres = float(i["Value"][j])
                if (thres < i["Threshold"][0]):
                    sheet[A3].font = Font(color= "00000000");
                    sheet[A3].fill =PatternFill(fill_type="solid",start_color='0000FF00') #Green
                elif(thres > i["Threshold"][0] and thres < i["Threshold"][1]):
                    sheet[A3].font = Font(color= "00000000");
                    sheet[A3].fill =PatternFill(fill_type="solid",start_color='00FFFF99'); #yellow
                elif (thres > i["Threshold"][1]):
                    sheet[A3].font = Font(color= "00000000");
                    sheet[A3].fill =PatternFill(fill_type="solid",start_color='00FF0000'); #RED

            
            # print(i["Value"]);


#code for updating the cell width value

for i in av:
    if i["subtitle"] != []:
        for k in i["subtitle"]:
            max_len = 0
            for l in k["Value"]:
                if max_len < len(l):
                    max_len = len(l)
            if max_len < len(k["title"]):
                max_len = len(k["title"])
            sheet.column_dimensions[f'{k["col"]}'].width = max_len+2;
        continue
    else:
        max_len = 0
        for j in i["Value"]:
            if max_len < len(j):
                max_len = len(j)
        if max_len < len(i["title"]):
            max_len = len(i["title"])
        sheet.column_dimensions[f'{i["col"]}'].width = max_len+2;
      
# print("The max row value is :",sheet.max_row)3095
# print("The maximun column value is:",sheet.max_column)

Congestion_Image = {
    "title": "Congestion_Image",
    "Value": [],
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

Snap_img_list = []

try:
    Snaps = f"ls 35_pnr/rpts/*/*.png"
    Snap_Img = subprocess.check_output(Snaps, shell=True).decode("utf-8")
    #print("try")
    Snap_img_list = Snap_Img.split("\n")
    if Snap_Img == "":
        Snap_img_list.append("")

except:
    Snap_img_list.append("")


#print("THe snap list is", Snap_img_list)
if len(Snap_img_list) == 1:
    print("NO, Image is Present")

sheet_no = []
total_stage_snap = [];
for i in range(0, len(Snap_img_list) - 1):
    a = list(Snap_img_list[i].split("/"))[2]
    total_stage_snap.append(a);

#print ("total stage without sorting is", total_stage_snap);    
total_stage_snap = list(OrderedDict.fromkeys(total_stage_snap))
#l = list(OrderedDict.fromkeys(l))
for i in range(0, len(total_stage_snap)):
    sheet_no.append(f"sheet{i+1}")

#print("the sheet_name is", sheet_no)
A1 = ["A1", "AF1", "A60", "AF60","A120", "AF120"];  #name of the image
A2 = ["A2", "AF2", "A61", "AF61","A121", "AF121"];   #location of the Image

X1 = ["F1", "AK1", "F60", "AK60","F120", "AK120"];  # for merging the cell, the last point until which the cells must geet merged 

print("the total stage which has images are :",total_stage_snap);


for i in range(0,len(total_stage_snap)):
  a = f"ls 35_pnr/rpts/{total_stage_snap[i]}/*.png";
  #print(a);
  stg_images = subprocess.run(a, shell=True, capture_output = True, text = True );
  #print ("the stage images possible are: ",stg_images.stdout);
  if (stg_images.stderr == ""):
    stg_img_ans = stg_images.stdout.strip().split();
    #print("types is ",type(stg_img_ans))
    #print("the stge img ans is", stg_img_ans)
    name1 = "35_pnr" + "__" + total_stage_snap[i]; 
    sheet_no[i] = work_obj.create_sheet(name1);
    if (len(stg_img_ans) > 6):
      print(f"YOU have more than 6 images in {stg_img_ans[0].split('/')[-2]} stage, only first 6 will be shown in the sheet")
    for j in range(0, min(len(stg_img_ans), len(A1))):
        sheet_no[i][f'{A1[j]}'] = stg_img_ans[j].split("/")[-1];
        sheet_no[i].merge_cells(f'{A1[j]}:{X1[j]}')
        sheet_no[i][f'{A1[j]}'].font = Congestion_Image["font"];
        sheet_no[i][f'{A1[j]}'].fill = Congestion_Image["fill"];
        sheet_no[i][f'{A1[j]}'].alignment = Congestion_Image["align"];
        sheet_no[i][f'{A1[j]}'].border = Congestion_Image["border"];
        img = Image(f'{sys.argv[1]}/{stg_img_ans[j]}');
        img.anchor = A2[j];
        sheet_no[i].add_image(img, img.anchor);
        
    




#Snap_img_list = []

#try:
    #Snaps = f"ls 35_pnr/rpts/*/*.png"
    #Snap_Img = subprocess.check_output(Snaps, shell=True).decode("utf-8")
    ##print("try")
    #Snap_img_list = Snap_Img.split("\n")
    #if Snap_Img == "":
        #Snap_img_list.append("")

#except:
    #Snap_img_list.append("")


##print("THe snap list is", Snap_img_list)
#if len(Snap_img_list) == 1:
    #print("NO, Image is Present")

#sheet_no = []
#for i in range(0, len(Snap_img_list) - 1):
    #sheet_no.append(f"sheet{i+1}")

##print("the sheet_name is", sheet_no)

#for i in range(0, len(Snap_img_list) - 1):
    #name = list(Snap_img_list[i].split("/"))[2]
    #name1 = list(Snap_img_list[i].split("/"))[3]
    #print("The name type is", type(name))
    #name = f'35_pnr_{name}'
    #sheet_no[i] = work_obj.create_sheet(name1)
    #sheet_no[i].merge_cells("A1:D1")
    #sheet_no[i]["A1"] = name
    #cong_img = f"{sys.argv[1]}/{Snap_img_list[i]}"
    #img = Image(cong_img)
    #img.anchor = "A2"
    #sheet_no[i].add_image(img, img.anchor)


print("\n\n")


#-----------------------------------------------------------------------------------------------------
#For_Checklist

sheet_checklist = work_obj.create_sheet("Task_Checklist");

STD_STRUCT_PRESERVE = {
    "title": "STD_STRUCT_PRESERVE in_icc2",
    "Value": stage_std_struct,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : ["PASS", "FAIL"],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
   
Preserve_ui = {
    "title": "PRESERVE UI",
    "Value": stage_preserve,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : ["PASS", "FAIL"],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

dp_flow = {
    "title": f'{file_des}_SVAR(dp,snap_grid,exclude_macro,lib_cell_pattern_list)',
    "Value": stage_dp_grid,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : ["PASS", "FAIL"],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
dp_check_stage_25 = {
    "title": "DP stage_name",
    "Value": check_stage_name,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

dp_icc_ver = {
    "title": "DP_icc_version(USED)",
    "Value": stage_icc_ver,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
dp_golden_ver = {
    "title": f'DP_ Golden_version',
    "Value": icc_golden_ver,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : [],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}
dp_icc_ver_status = {
    "title": "DP icc ver_ status",
    "Value": stage_icc_ver_status,
    "Width": "NA",
    "Row": [],
    "col": "NA",
    "font": Font(bold=True, color="00000000"),
    "Threshold" : ["PASS", "FAIL"],
    "fill": PatternFill(fill_type="solid", start_color="00FFFF00"),
    "align": Alignment(horizontal="center", vertical="center"),
    "border": Border(
        left=Side(border_style="medium", color="00333300"),
        right=Side(border_style="medium", color="00333300"),
        top=Side(border_style="medium", color="00333300"),
        bottom=Side(border_style="medium", color="00333300"),
    ),
    "subtitle": [],
}

checklist = [
     STD_STRUCT_PRESERVE,
     Preserve_ui,
     dp_flow,
  ]

av_dp = [
    dp_check_stage_25,
    dp_icc_ver_status,
    dp_icc_ver,
    dp_golden_ver,
]


###------------------------------------------------------------------------------------------------
for dv in checklist:
    if (dv == dp_flow):
        sheet_checklist.append([dv["title"], dv["Value"][0]]);
    else:
        sheet_checklist.append([dv["title"], dv["Value"][0]]);

#check_gap= 3;
w = sheet_checklist.max_row;
x =w;

for i in ['A', 'B']:
    max_val = 0;
    for j in range(1, x +1):
        sheet_checklist[f'A{j}'].fill = title["fill"];
        sheet_checklist[f'A{j}'].border = title["border"];
        if (sheet_checklist[f'B{j}'].value == "PASS"):
            sheet_checklist[f'B{j}'].fill = PatternFill(fill_type="solid",start_color='0000FF00');
        else:
            sheet_checklist[f'B{j}'].fill = PatternFill(fill_type="solid",start_color='00FF0000');
        if (len(sheet_checklist[f'{i}{j}'].value)!=0 and len(sheet_checklist[f'{i}{j}'].value) > max_val):
            max_val = len(sheet_checklist[f'{i}{j}'].value);
    sheet_checklist.column_dimensions[i].width = max_val +1;
    
#25_DP additon

#for i in range(w, w+ check_gap):
    #sheet_checklist.append([""]);

#no_title = False;

#title_check = [];
#for i in av_dp:
    #title_check.append(i["title"]);

#sheet_checklist.append(title_check);

##a= sheet_checklist.max_row;
##b = get_column_letter(sheet_checklist.max_column);
##print("the value a",a);
##print("the value b", b);
##sheet_checklist[f'{a}{b}'].fill = PatternFill(fill_type="solid", start_color="00FFFF00");
###sheet_checklist[f'{a}{b}'].font = Font();

#for j in range(0,len(av_dp[0]["Value"])):
    #check_dp = [];
    
    #for i in av_dp:
        #check_dp.append(i["Value"][j]);
    #sheet_checklist.append(check_dp);
   
                   






# for i in range(0, len(checklist)):
#     checklist[i]["col"] = chr(65 + i)
#     checklist[i]["Width"] = len(checklist[i]["title"]) + 2
#     # Changing the width of the column based on the title length


# row_li = [1, 2]

#print("The value of xis",x)
row_li = [];
for row in range(x+2, x+2 +2):
    row_li.append(row);

#print("The row li is",row_li);

for i in range(row_li[-1]+1, len(check_stage_name) + row_li[-1]+1):
    row_li.append(i)
#print("The row li is :")
#print(row_li)

for C in av_dp:
    if C["subtitle"] == []:
        C["Row"] = row_li
    else:
        C["Row"] = row_li
        for c in C["subtitle"]:
            c["Row"] = row_li


# Code for Assigning col value to each dictionary
is_array = True
i = 0
k = 0
while is_array:
    if k == len(av_dp) - 1:
        is_array = False

    if av_dp[k]["subtitle"] == []:
        av_dp[k]["col"] = chr(65 + i)
        i = i + 1
    else:
        av_dp[k]["col"] = chr(65 + i)

        j = 0
        for sub_title in av_dp[k]["subtitle"]:
            sub_title["col"] = chr(65 + i + j)
            j = j + 1
        i = i + j
    # temp= av_dp[k]["title"]
    k = k + 1


# code for merging the desired cells
for dic in av_dp:
    if dic["subtitle"] == []:
        ans = f'{dic["col"]}{dic["Row"][1]}:{dic["col"]}{dic["Row"][1]}'
        #print ("the ans", ans);
        sheet_checklist.merge_cells(ans)
    else:
        ans = f'{dic["subtitle"][0]["col"]}1:{dic["subtitle"][-1]["col"]}1'
        # print(ans, type(ans), end= " ")
        sheet_checklist.merge_cells(ans)

print("YYYY")

# code for adding Title , value,font, fill, align,
for title in av_dp:
    if title["subtitle"] == []:
        tit = f'{title["col"]}{title["Row"][1]}'
        # print("the title is")
        # print(title["col"], title["Row"][0]);
        sheet_checklist[tit] = title["title"]

        sheet_checklist[tit].font = title["font"]
        sheet_checklist[tit].fill = title["fill"]
        sheet_checklist[tit].alignment = title["align"]
        sheet_checklist[tit].border = title["border"]

        # value = f'{title["col"]}{title["Row"][-1]}'
        # sheet_checklist[value] = title["Value"];
    else:
        #print(title["col"], title["Row"][0])
        tit = f'{title["col"]}{title["Row"][0]}'
        #print("the SUB_title is")
        #print(title["col"], title["Row"][0])
        sheet_checklist[tit] = title["title"]
        sheet_checklist[tit].font = title["font"]
        sheet_checklist[tit].fill = title["fill"]
        sheet_checklist[tit].alignment = title["align"]
        sheet_checklist[tit].border = title["border"]
        for sub in title["subtitle"]:
            tit = f'{sub["col"]}{sub["Row"][1]}'
            sheet_checklist[tit] = sub["title"]

            sheet_checklist[tit].font = sub["font"]
            sheet_checklist[tit].fill = sub["fill"]
            sheet_checklist[tit].alignment = sub["align"]
            sheet_checklist[tit].border = sub["border"]
            # value = f'{sub["col"]}{sub["Row"][-1]}'
            # sheet_checklist[value] = sub["Value"];

# inserting values

for i in av_dp:
    if i["Value"] == []:
        if i["subtitle"] == []:
            continue
        else:
            for j in i["subtitle"]:
                if j["Value"] == []:
                    continue
                else:
                    for k in range(0, len(j["Value"])):
                        A3 = f'{j["col"]}{j["Row"][k+2]}'
                        sheet_checklist[A3].value = j["Value"][k]
    else:
        for j in range(0, len(i["Value"])):
            A3 = f'{i["col"]}{i["Row"][j+2]}'
            sheet_checklist[A3].value = i["Value"][j]
            if (i["Threshold"]  != []):
                thres = (i["Value"][j])
                if (thres == i["Threshold"][0]):
                    sheet_checklist[A3].font = Font(color= "00000000");
                    sheet_checklist[A3].fill =PatternFill(fill_type="solid",start_color='0000FF00') #Green
                elif(thres > i["Threshold"][0] and thres < i["Threshold"][1]):
                    sheet_checklist[A3].font = Font(color= "00000000");
                    sheet_checklist[A3].fill =PatternFill(fill_type="solid",start_color='00FFFF99'); #yellow
                elif (thres == i["Threshold"][1]):
                    sheet_checklist[A3].font = Font(color= "00000000");
                    sheet_checklist[A3].fill =PatternFill(fill_type="solid",start_color='00FF0000'); #RED
   
            # print(i["Value"]);


#code for updating the cell width value

for i in av_dp:
    if i["subtitle"] != []:
        for k in i["subtitle"]:
            max_len = 0
            for l in k["Value"]:
                if max_len < len(l):
                    max_len = len(l)
            if max_len < len(k["title"]):
                max_len = len(k["title"])
            sheet_checklist.column_dimensions[f'{k["col"]}'].width = max_len+1;
        continue
    else:
        max_len = 0
        for j in i["Value"]:
            if max_len < len(j):
                max_len = len(j)
        if max_len < len(i["title"]):
            max_len = len(i["title"])
        sheet_checklist.column_dimensions[f'{i["col"]}'].width = max_len+1;
      
# print("The max row value for check_list is :",sheet_checklist.max_row)
# print("The maximun column value check-lsit is:",sheet_checklist.max_column)



#---------------------------------------------------------------------------------------------
#print(curr_dir);
#os.system(f"cd {ppp2_file}")
os.chdir(ppp2_file);

os.system("pwd");
work_obj.save(f"{ppp2_file}/{ppp2}_{design_name}.xls")

#os.system(f"soffice {ppp2_file}/{ppp2}_{design_name}.xls")
